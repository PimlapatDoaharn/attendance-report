from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys
import tempfile
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from attendance_report.processor import ProcessingOptions, process_attendance_report


def create_finger_scan(path: Path) -> None:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "info (used)"
    ws.append(["Department", "ชื่อพนักงาน", "หมายเลขพนักงานบนระบบ", "วันและเวลา"])
    ws.append(["Sales", "Somchai Ja", "E999", "05/05/2026 08:01"])
    ws.append(["Sales", "Somchai Jai", "E001", "05/05/2026 18:02"])
    ws.append(["Sales", "Suda Dee", "E002", "Tue, May 6, 2026 08:11"])
    ws.append(["Sales", "Suda Dee", "E002", "07 พ.ค. 2569 08:11"])
    ws.append(["Sales", "Suda Dee", "E002", "09/05/2026 08:11"])
    ws.append(["Sales", "Suda Dee", "E002", "PM"])
    wb.save(path)


def create_report(path: Path) -> None:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "2026 Report"
    ws.append(["", "May", "May", "May"])
    ws.append(["", "Working Day", "Leave / Off-site work", "Target"])
    ws.append(["", 21, "", ""])
    ws.append(["ชื่อพนักงาน", "Working Day", "Leave / Off-site work", "Target", "Org Count"])
    ws.append(["Somchai Jai", "", "", "", "Strategy & Transformation (AI)"])
    ws.append(["Suda Dee", "", "", "", "Strategy & Transformation (AI)"])

    count = wb.create_sheet("Count scan")
    count.append(["placeholder"])

    leave = wb.create_sheet("ข้อมูลวันลา")
    leave.append(["ชื่อพนักงาน", "วันที่", "จำนวนวัน"])
    leave.append(["Somchai Jai", "06/05/2026", 0.5])

    hr = wb.create_sheet("ฐานข้อมูลพนักงาน (HR)")
    hr.append(["ชื่อพนักงาน", "หมายเลขพนักงานบนระบบ", "Department"])
    hr.append(["Somchai Jai", "E001", "Sales"])
    hr.append(["Suda Dee", "E002", "Sales"])

    prev = wb.create_sheet("Previous month")
    prev.append(["ชื่อพนักงาน", "หมายเลขพนักงานบนระบบ"])
    prev.append(["Somchai Jai", "E001"])
    wb.save(path)


def create_ptvn_report(path: Path) -> None:
    wb = Workbook()
    summary = cast(Worksheet, wb.active)
    summary.title = "2026 Summary Report"
    summary.append([None, "Departments", "NO employee", "Jan", "Feb", "Mar", "Apr", "May"])
    summary.append([None, None, None, None, None, None, None, None])
    summary.append([None, None, None, None, None, None, None, None])
    summary.append([None, None, None, None, None, None, None, None])
    summary.append([None, "Strategy & Transformation (AI)", None, None, None, None, None, None])
    summary.append([None, "PTV Overall", None, None, None, None, None, None])

    individual = wb.create_sheet("2026 Individual Attendance Repo")
    individual.append(["", "May", "May", "", "May"])
    individual.append(["", "Working Day", "Leave / Off-site work", "Org Count", "Target 60%"])
    individual.append(["", "", "", "", ""])
    individual.append(["ชื่อพนักงาน", "Working Day", "Leave / Off-site work", "Org Count", "Target 60%"])
    individual.append(["Somchai Jai", "", "", "Strategy & Transformation (AI)", ""])
    individual.append(["Suda Dee", "", "", "Strategy & Transformation (AI)", ""])
    wb.save(path)


def main() -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)
        finger = temp / "finger.xlsx"
        report = temp / "report.xlsx"
        ptvn = temp / "ptvn.xlsx"
        output = temp / "output.xlsx"
        ptvn_output = temp / "ptvn_output.xlsx"
        create_finger_scan(finger)
        create_report(report)
        create_ptvn_report(ptvn)

        process_attendance_report(
            finger,
            report,
            output,
            ProcessingOptions(month=5, year=2026),
            ptvn_report_path=ptvn,
            ptvn_output_path=ptvn_output,
            log=print,
        )

        wb = load_workbook(output, data_only=False)
        ptvn_wb = load_workbook(ptvn_output, data_only=False)
        assert "Raw finger scan" in wb.sheetnames
        assert "Count scan" in wb.sheetnames
        raw = wb["Raw finger scan"]
        count = wb["Count scan"]
        report_sheet = wb["2026 Report"]
        assert raw.max_row == 7
        assert raw["E7"].value is None
        assert raw["F7"].value is None
        assert raw["E2"].number_format == "dd/mm/yyyy"
        assert raw["F2"].number_format == "ddd,mmm d,yyyy"
        assert str(raw["E2"].value).startswith("2026-05-05")
        assert str(raw["F2"].value).startswith("2026-05-05")
        assert count["A2"].value == "COUNTUNIQUE of ชื่อพนักงาน"
        header_values = [cell.value for cell in count[3]]
        scan_dates = [value for value in header_values if isinstance(value, datetime)]
        assert scan_dates
        assert all(value.weekday() < 5 for value in scan_dates)
        grand_total_col = header_values.index("Grand Total") + 1
        total_col = header_values.index("Total") + 1
        assert count.cell(3, grand_total_col).value == "Grand Total"
        assert str(count.cell(5, total_col).value).startswith("=sum(C5:")
        assert report_sheet["B3"].value == 21
        assert report_sheet["D3"].value == 13
        assert report_sheet["B5"].value == "=IFERROR(VLOOKUP(A5,'Count Scan'!A:G,7,0),0)"
        assert report_sheet["C5"].value == "=SUMIF('ข้อมูลวันลา'!B:B, A5, 'ข้อมูลวันลา'!J:J)"
        assert report_sheet["D5"].value == "=(B5+C5)/$B$3"
        summary = ptvn_wb["2026 Summary Report"]
        individual = ptvn_wb["2026 Individual Attendance Repo"]
        dashboard = ptvn_wb["Dashboard"]
        assert str(summary["C5"].value).startswith("=COUNTIF('2026 Individual Attendance Repo'!$D:$D")
        assert summary["H3"].value == 21
        assert str(summary["H5"].value).startswith("=IFERROR(ROUNDUP(AVERAGEIF('2026 Individual Attendance Repo'!$D:$D")
        assert summary["H5"].number_format == "0%"
        assert str(summary["C6"].value).startswith("=SUM(C")
        assert str(summary["H6"].value).startswith("=AVERAGE(H")
        assert summary["H6"].number_format == "0%"
        assert individual["B3"].value == 21
        assert individual["E3"].value == 13
        assert individual["B5"].value == 1
        assert individual["C5"].value == 1
        assert individual["E5"].value == "=MIN(ROUNDUP((B5+C5)/$B$3,2),100%)"
        assert individual["E5"].number_format == "0%"
        assert "PTVN Attendance Dashboard" in str(dashboard["C1"].value)
        assert "ตัวกรองข้อมูล" in str(dashboard["A4"].value)
        assert dashboard["A5"].value.startswith("เดือน")
        assert dashboard["A11"].value.startswith("แผนก")
        assert dashboard["A18"].value.startswith("พนักงาน")
        assert dashboard["A7"].value is None
        assert all(dashboard[f"A{row}"].value is None for row in range(13, 18))
        assert len(dashboard.data_validations.dataValidation) == 3
        validation_ranges = [str(validation.sqref) for validation in dashboard.data_validations.dataValidation]
        assert validation_ranges == ["A6", "A12", "A19"]
        assert dashboard.column_dimensions["AH"].hidden
        assert "Working Days" in str(dashboard["E4"].value)
        assert "PTV Overall" in str(dashboard["G4"].value)
        assert str(dashboard["C4"].value).startswith('="Total Employees"&CHAR(10)&IF($A$19<>')
        assert str(dashboard["C91"].value).startswith('=IF($A$12="ทั้งหมด"')
        assert str(dashboard["D91"].value).startswith('=IF($A$12="ทั้งหมด"')
        assert str(dashboard["E91"].value).startswith('=IF($A$12="ทั้งหมด"')
        assert dashboard["E91"].number_format == "0%"
        assert str(dashboard["ED14"].value).startswith('="Pass"&CHAR(10)&EE14')
        assert str(dashboard["EE14"].value).startswith("=IFERROR(SUMPRODUCT")
        assert str(dashboard["ED15"].value).startswith('="Below Target"&CHAR(10)&EE15')
        assert str(dashboard["EE15"].value).startswith("=IFERROR(SUMPRODUCT")
        assert dashboard["EG14"].value in {"Jan", "May"}
        assert "$A$19" in str(dashboard["EH14"].value)
        assert "$A$12" in str(dashboard["EH14"].value)
        assert dashboard["EI14"].value == "=60%"
        assert dashboard["AN14"].value == "Strategy & Transformation (AI)"
        assert dashboard["AO14"].value == "='2026 Summary Report'!C5"
        assert str(dashboard["AP14"].value).startswith("='2026 Summary Report'!")
        assert dashboard["BD2"].value == "Somchai Jai"
        assert dashboard["BE2"].value == "Strategy & Transformation (AI)"
        assert str(dashboard["BF2"].value).startswith("=IFERROR(INDEX($BJ2:")
        assert dashboard["BG2"].value == "strategytransformationai"
        assert "Emp_" in str(dashboard["BH2"].value)
        assert "INDEX($BI$14:$BI$" in str(dashboard["BH3"].value)
        assert dashboard["BI14"].value == "strategytransformationai"
        assert "Emp_All" in ptvn_wb.defined_names
        assert "Emp_strategytransformationai" in ptvn_wb.defined_names
        employee_validation = dashboard.data_validations.dataValidation[2]
        assert employee_validation.formula1 == "=INDIRECT($BH$2)"
        assert "Pass Employee" in str(dashboard["M4"].value)
        assert "Below Target Employee" in str(dashboard["O4"].value)
        assert str(dashboard["M4"].value).startswith('="Pass Employee"&CHAR(10)&IFERROR(SUMPRODUCT')
        assert "$A$19" in str(dashboard["M4"].value)
        assert str(dashboard["A20"].value).startswith("=IFERROR(INDEX(INDIRECT($BH$2)")
        assert len(cast(Any, dashboard)._charts) == 5
        assert dashboard["EJ40"].value == "Other"
        assert dashboard["EJ41"].value == "Sick Leave"
        assert str(dashboard["EK40"].value).startswith("=IFERROR(SUMPRODUCT")
        assert dashboard["EP40"].value == "Total Absence Days"
        assert str(dashboard["EQ40"].value).startswith("=IFERROR(SUMPRODUCT")
        assert str(dashboard["EJ14"].value) == "\u226580%"
        assert str(dashboard["EJ18"].value) == "<20%"
        assert str(dashboard["EK14"].value).startswith("=IFERROR(SUMPRODUCT")
        chart_anchors = [(chart.anchor._from.col, chart.anchor._from.row) for chart in cast(Any, dashboard)._charts]
        assert chart_anchors == [(2, 9), (7, 9), (2, 46), (2, 28), (2, 66)]
        print(f"Smoke test passed: {output}")


if __name__ == "__main__":
    main()

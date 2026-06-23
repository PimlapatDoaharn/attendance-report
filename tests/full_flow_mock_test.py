from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys
import tempfile
from typing import Any, cast
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from attendance_report.processor import ProcessingOptions, process_attendance_report


EMPLOYEES = [
    ("Varanyu Visansirikul", "E001", "Strategy & Transformation (AI)", 5),
    ("Supavit Jaraspornsrivong", "E002", "Strategy & Transformation (AI)", 4),
    ("Nisra Bunnag", "E003", "Dev (AI-Augmented)", 5),
    ("Chananan Saisiri", "E004", "Dev (AI-Augmented)", 2),
    ("Siriruch Leelerdpong", "E005", "Customer Success", 3),
    ("Atchara Damngarm", "E006", "Customer Success", 1),
    ("Petchkarat Manowisut", "E007", "Auction Management", 4),
    ("Kunpriya Monthianthong", "E008", "Auction Management", 0),
    ("Patcharaphan Auarattanasakulchai", "E009", "HR", 5),
    ("Yatika Ngampriam", "E010", "HR", 3),
]

DEPARTMENTS = [
    "Strategy & Transformation (AI)",
    "Dev (AI-Augmented)",
    "Customer Success",
    "Auction Management",
    "HR",
]

SCAN_DATES = [
    "04/05/2026 08:01",
    "05/05/2026 08:02",
    "06/05/2026 08:03",
    "07/05/2026 08:04",
    "08/05/2026 08:05",
]


def create_full_finger_scan(path: Path) -> None:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "info (used)"
    ws.append(["Department", "ชื่อพนักงาน", "หมายเลขพนักงานบนระบบ", "วันและเวลา"])
    for name, employee_id, department, days_present in EMPLOYEES:
        for scan_datetime in SCAN_DATES[:days_present]:
            ws.append([department, name, employee_id, scan_datetime])
        if days_present:
            ws.append([department, name, employee_id, SCAN_DATES[0]])
    ws.append(["Unknown", "Not In HR", "EXXX", "04/05/2026 08:00"])
    ws.append(["HR", "Yatika Ngampriam", "E010", "PM"])
    ws.append(["HR", "Yatika Ngampriam", "E010", "04/04/2026 08:00"])
    wb.save(path)


def create_full_report_template(path: Path) -> None:
    wb = Workbook()
    report = cast(Worksheet, wb.active)
    report.title = "2026 Report"
    report.append(["", "May", "May", "May", ""])
    report.append(["", "Working Day", "Leave / Off-site work", "Target", ""])
    report.append(["", "", "", "", ""])
    report.append(["ชื่อพนักงาน", "Working Day", "Leave / Off-site work", "Target", "Org Count"])
    for name, _, department, _ in EMPLOYEES:
        report.append([name, "", "", "", department])

    count = wb.create_sheet("Count scan")
    count.append(["existing"])

    leave = wb.create_sheet("ข้อมูลวันลา")
    leave.append(["ชื่อพนักงาน", "วันที่", "จำนวนวัน"])
    leave.append(["Chananan Saisiri", "06/05/2026", 1])
    leave.append(["Atchara Damngarm", "07/05/2026", 0.5])
    leave.append(["Kunpriya Monthianthong", "07/05/2026", 2])
    leave.append(["Varanyu Visansirikul", "04/04/2026", 1])

    hr = wb.create_sheet("ฐานข้อมูลพนักงาน (HR)")
    hr.append(["ชื่อพนักงาน", "หมายเลขพนักงานบนระบบ", "Department"])
    for name, employee_id, department, _ in EMPLOYEES:
        hr.append([name, employee_id, department])

    previous = wb.create_sheet("Previous month")
    previous.append(["ชื่อพนักงาน", "หมายเลขพนักงานบนระบบ"])
    for name, employee_id, _, _ in EMPLOYEES:
        previous.append([name, employee_id])

    wb.save(path)


def create_full_ptvn_template(path: Path) -> None:
    wb = Workbook()
    summary = cast(Worksheet, wb.active)
    summary.title = "2026 Summary Report"
    summary.append([None, "Departments", "NO employee", "Mar", "Apr", "May", "Jun", "Jul"])
    summary.append([None, None, None, None, None, None, None, None])
    summary.append([None, None, None, 18, 19, None, 0, 0])
    summary.append([None, None, None, None, None, None, None, None])
    summary.append([None, "Target", None, "60%", "60%", "60%", "60%", "60%"])
    for index, department in enumerate(DEPARTMENTS, start=6):
        summary.append([None, department, len([employee for employee in EMPLOYEES if employee[2] == department]), 0.55 + index / 100, 0.50 + index / 100, None, 0, 0])
    summary.append([None, "PTV Overall", len(EMPLOYEES), 0.62, 0.58, None, 0, 0])
    summary.merge_cells("B5:C5")

    individual = wb.create_sheet("2026 Individual Attendance Repo")
    individual.append(["Employees are now required to work in-person at least twice a week", None, None, None, None])
    individual.append(["", "", "May", "", "May", "May"])
    individual.append(["", "", "Working Day", "", "Leave / Off-site work", "Target 60%"])
    individual.append(["No.", "Name", "Working Day", "Org Count", "Leave / Off-site work", "Target 60%"])
    for index, (name, _, department, _) in enumerate(EMPLOYEES, start=1):
        individual.append([index, name, "", department, "", ""])
    individual.merge_cells("A1:F2")

    wb.save(path)


def assert_no_chart_overlap(dashboard: Worksheet) -> None:
    anchors = []
    for chart in cast(Any, dashboard)._charts:
        marker = chart.anchor._from
        anchors.append((marker.col, marker.row))
    assert anchors == [(2, 8), (9, 8), (2, 28)]
    assert len(set(anchors)) == len(anchors)


def main() -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)
        finger = temp / "finger_full.xlsx"
        report = temp / "report_full.xlsx"
        ptvn = temp / "ptvn_full.xlsx"
        output = temp / "output_full.xlsx"
        ptvn_output = temp / "ptvn_output_full.xlsx"

        create_full_finger_scan(finger)
        create_full_report_template(report)
        create_full_ptvn_template(ptvn)

        logs: list[str] = []
        process_attendance_report(
            finger,
            report,
            output,
            ProcessingOptions(month=5, year=2026, working_days=5),
            ptvn_report_path=ptvn,
            ptvn_output_path=ptvn_output,
            log=logs.append,
        )

        assert any("skipped 1 rows with no HR match" in message for message in logs)
        assert any("skipped 1 rows with invalid scan date/time" in message for message in logs)

        wb = load_workbook(output, data_only=False)
        ptvn_wb = load_workbook(ptvn_output, data_only=False)

        count = wb["Count scan"]
        report_sheet = wb["2026 Report"]
        assert count.max_row >= len(EMPLOYEES) + 5
        assert report_sheet["B3"].value == 5
        assert report_sheet["D3"].value == 3
        assert report_sheet["B5"].value == "=IFERROR(VLOOKUP(A5,'Count Scan'!A:I,9,0),0)"
        assert report_sheet["C8"].value == "=SUMIF('ข้อมูลวันลา'!B:B, A8, 'ข้อมูลวันลา'!J:J)"
        assert report_sheet["D8"].value == "=(B8+C8)/$B$3"

        summary = ptvn_wb["2026 Summary Report"]
        individual = ptvn_wb["2026 Individual Attendance Repo"]
        dashboard = ptvn_wb["Dashboard"]

        assert summary["B5"].value == "Target"
        assert summary["C5"].value is None
        assert summary["F3"].value == 5
        assert str(summary["C6"].value).startswith("=COUNTIF('2026 Individual Attendance Repo'!$D:$D")
        assert str(summary["F6"].value).startswith("=IFERROR(ROUNDUP(AVERAGEIF('2026 Individual Attendance Repo'!$D:$D")
        assert str(summary["C11"].value).startswith("=SUM(C6:C10)")
        assert str(summary["F11"].value).startswith("=AVERAGE(F6:F10)")

        assert individual["C3"].value == 5
        assert individual["F3"].value == 3
        assert individual["C5"].value == 5
        assert individual["F5"].value == "=MIN(ROUNDUP((C5+E5)/$C$3,2),100%)"
        assert individual["F8"].value == "=MIN(ROUNDUP((C8+E8)/$C$3,2),100%)"

        assert dashboard["A6"].value == "พฤษภาคม 2569"
        assert dashboard["A12"].value == "ทั้งหมด"
        assert dashboard["A19"].value == "ทั้งหมด"
        assert len(dashboard.data_validations.dataValidation) == 3
        assert dashboard.data_validations.dataValidation[2].formula1 == "=INDIRECT($BH$2)"
        assert "Emp_All" in ptvn_wb.defined_names
        assert "Emp_strategytransformationai" in ptvn_wb.defined_names
        assert "Emp_devaiaugmented" in ptvn_wb.defined_names
        assert dashboard["BD2"].value == "Varanyu Visansirikul"
        assert dashboard["BE2"].value == "Strategy & Transformation (AI)"
        assert dashboard["BD4"].value == "Nisra Bunnag"
        assert dashboard["BE4"].value == "Dev (AI-Augmented)"
        assert str(dashboard["BH2"].value).startswith('=IF($A$12="ทั้งหมด","Emp_All"')
        assert str(dashboard["BH3"].value).startswith('=IF($A$12="ทั้งหมด","ทั้งหมด"')

        assert str(dashboard["C4"].value).startswith('="Total Employees"&CHAR(10)&IF($A$19<>')
        assert str(dashboard["G4"].value).startswith('="PTV Overall"&CHAR(10)&TEXT(')
        assert str(dashboard["M4"].value).startswith('="Pass Employee"&CHAR(10)&IFERROR(SUMPRODUCT')
        assert str(dashboard["O4"].value).startswith('="Below Target Employee"&CHAR(10)&IFERROR(SUMPRODUCT')
        assert "$A$19" in str(dashboard["M4"].value)
        assert dashboard["W13"].value == "Department"
        assert dashboard["X13"].value == "Attendance"
        assert dashboard["Y13"].value is None
        assert str(dashboard["Z14"].value).startswith('="Pass"&CHAR(10)&AA14')
        assert str(dashboard["Z15"].value).startswith('="Below Target"&CHAR(10)&AA15')
        assert dashboard["AC14"].value == "Mar"
        assert dashboard["AC15"].value == "Apr"
        assert dashboard["AH4"].value == "พฤษภาคม 2569"
        assert str(dashboard["AN2"].value).startswith("=COUNTA('2026 Individual Attendance Repo'!$B$5:$B")
        assert dashboard["AN14"].value == "Strategy & Transformation (AI)"
        assert dashboard["AO14"].value == "='2026 Summary Report'!C6"
        assert str(dashboard["AP14"].value).startswith("='2026 Summary Report'!D6")
        assert str(dashboard["AR14"].value).startswith("='2026 Summary Report'!F6")
        assert dashboard["AF40"].value == "Other"
        assert dashboard["AF41"].value == "Sick Leave"
        assert str(dashboard["AG40"].value).startswith("=IFERROR(SUMPRODUCT")
        assert dashboard["AL40"].value == "Total Absence Days"
        assert str(dashboard["AM41"].value).startswith("=IFERROR(")
        assert str(dashboard["A20"].value).startswith("=IFERROR(INDEX(INDIRECT($BH$2)")

        assert_no_chart_overlap(dashboard)
        assert dashboard.sheet_view.showGridLines is False
        assert dashboard.column_dimensions["AH"].hidden

        with zipfile.ZipFile(ptvn_output) as archive:
            chart_files = [name for name in archive.namelist() if name.startswith("xl/charts/chart")]
            drawing_files = [name for name in archive.namelist() if name.startswith("xl/drawings/drawing")]
            dashboard_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
            chart_xml = "\n".join(archive.read(name).decode("utf-8") for name in chart_files)
        assert len(chart_files) == 3
        assert drawing_files
        assert "FILTER" not in dashboard_xml
        assert "SORTBY" not in dashboard_xml
        assert "_xlfn" not in dashboard_xml
        assert "<c:dLbls" not in chart_xml
        assert dashboard["J35"].value.startswith("No absence category data available")

        ptvn_xlsm = temp / "ptvn_template_full.xlsm"
        ptvn_xlsm_output = temp / "ptvn_output_full.xlsm"
        create_full_ptvn_template(ptvn_xlsm)
        process_attendance_report(
            finger,
            report,
            output,
            ProcessingOptions(month=5, year=2026, working_days=19),
            ptvn_report_path=ptvn_xlsm,
            ptvn_output_path=ptvn_xlsm_output,
            log=lambda _message: None,
        )
        ptvn_xlsm_wb = load_workbook(ptvn_xlsm_output, keep_vba=True, data_only=False)
        xlsm_dashboard = ptvn_xlsm_wb["Dashboard"]
        assert len(cast(Any, xlsm_dashboard)._charts) == 0
        with zipfile.ZipFile(ptvn_xlsm_output) as archive:
            assert not [name for name in archive.namelist() if name.startswith("xl/charts/chart")]

        print(f"Full mock flow test passed: {output}")
        print(f"Full mock PTVN output passed: {ptvn_output}")


if __name__ == "__main__":
    main()

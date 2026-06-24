from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
import math
import re
from pathlib import Path
from typing import Any, Callable, Iterable, Optional, cast

from dateutil import parser as date_parser
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz, process

LogFn = Callable[[str], None]

THAI_MONTHS = {
    "ม.ค.": 1,
    "มกราคม": 1,
    "ก.พ.": 2,
    "กุมภาพันธ์": 2,
    "มี.ค.": 3,
    "มีนาคม": 3,
    "เม.ย.": 4,
    "เมษายน": 4,
    "พ.ค.": 5,
    "พฤษภาคม": 5,
    "มิ.ย.": 6,
    "มิถุนายน": 6,
    "ก.ค.": 7,
    "กรกฎาคม": 7,
    "ส.ค.": 8,
    "สิงหาคม": 8,
    "ก.ย.": 9,
    "กันยายน": 9,
    "ต.ค.": 10,
    "ตุลาคม": 10,
    "พ.ย.": 11,
    "พฤศจิกายน": 11,
    "ธ.ค.": 12,
    "ธันวาคม": 12,
}

WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
THAI_MONTH_NAMES = [
    "มกราคม",
    "กุมภาพันธ์",
    "มีนาคม",
    "เมษายน",
    "พฤษภาคม",
    "มิถุนายน",
    "กรกฎาคม",
    "สิงหาคม",
    "กันยายน",
    "ตุลาคม",
    "พฤศจิกายน",
    "ธันวาคม",
]


@dataclass(frozen=True)
class Employee:
    name: str
    normalized_name: str
    employee_id: str
    department: str


@dataclass(frozen=True)
class FingerScanRow:
    source_department: str
    raw_name: str
    matched_name: str
    raw_employee_id: str
    resolved_employee_id: str
    raw_datetime: object
    scan_date: date
    hr_department: str


@dataclass(frozen=True)
class ProcessingOptions:
    month: int
    year: int
    fuzzy_threshold: int = 82
    working_days: int | None = None


def process_attendance_report(
    finger_scan_path: str | Path,
    attendance_report_path: str | Path,
    output_path: str | Path,
    options: ProcessingOptions,
    ptvn_report_path: str | Path | None = None,
    ptvn_output_path: str | Path | None = None,
    log: LogFn | None = None,
) -> Path:
    log = log or (lambda message: None)
    finger_scan_path = Path(finger_scan_path)
    attendance_report_path = Path(attendance_report_path)
    output_path = Path(output_path)

    log("Loading workbooks...")
    finger_wb = load_workbook(finger_scan_path, data_only=True)
    report_wb = load_workbook(attendance_report_path, keep_vba=attendance_report_path.suffix.lower() == ".xlsm")

    hr_sheet = require_sheet(report_wb, "ฐานข้อมูลพนักงาน (HR)")
    previous_month_sheet = find_sheet(report_wb, "Previous month")
    leave_sheet = find_sheet(report_wb, "ข้อมูลวันลา")
    report_sheet = require_sheet(report_wb, "2026 Report")

    employees = read_hr_employees(hr_sheet)
    if not employees:
        raise ValueError("ไม่พบข้อมูลพนักงานในชีท ฐานข้อมูลพนักงาน (HR)")
    log(f"Loaded {len(employees)} employees from HR sheet.")

    finger_rows_raw = read_finger_scan_rows(finger_wb)
    if not finger_rows_raw:
        raise ValueError("ไม่พบข้อมูลใน Report Finger Scan")
    log(f"Loaded {len(finger_rows_raw)} raw finger scan rows.")

    id_hint_by_name = read_previous_month_id_hints(previous_month_sheet)
    name_by_id = read_previous_month_name_by_id(previous_month_sheet)
    resolved_rows = normalize_finger_rows(
        finger_rows_raw,
        employees,
        id_hint_by_name,
        options,
        log,
        name_by_id=name_by_id,
    )
    log(f"Matched {len(resolved_rows)} finger scan rows.")
    total_raw = len(finger_rows_raw)
    has_name_date = sum(1 for r in finger_rows_raw if r.get("raw_name") and r.get("raw_datetime"))
    blank_rows = total_raw - has_name_date
    unmatched_hr = has_name_date - len(resolved_rows)
    log(
        f"Summary: {total_raw} total rows loaded\n"
        f"  ✅ Matched (HR):    {len(resolved_rows):>5}\n"
        f"  ❌ Unmatched (HR):  {unmatched_hr:>5}\n"
        f"  ⬜ Blank/no-date:   {blank_rows:>5}\n"
        f"  Match rate: {len(resolved_rows)/has_name_date*100:.1f}% of rows with name+date"
        if has_name_date else ""
    )

    write_raw_finger_scan_sheet(report_wb, finger_rows_raw, options)
    count_totals, count_scan_total_col = write_count_scan_sheet(report_wb, employees, finger_rows_raw, id_hint_by_name, options)
    log("Updated Raw finger scan and Count scan sheets.")

    leave_totals = read_leave_totals(leave_sheet, options) if leave_sheet else {}
    update_annual_report(report_sheet, count_totals, leave_totals, options, count_scan_total_col)
    log("Updated 2026 Report sheet.")

    if ptvn_report_path:
        ptvn_output = Path(ptvn_output_path) if ptvn_output_path else Path(ptvn_report_path)
        process_ptvn_report(
            ptvn_report_path,
            ptvn_output,
            report_sheet,
            employees,
            count_totals,
            leave_totals,
            options,
        )
        log(f"Updated PTVN report: {ptvn_output}")

    ensure_sheet_order(
        report_wb,
        [
            "2026 Report",
            "Count scan",
            "ข้อมูลวันลา",
            "ฐานข้อมูลพนักงาน (HR)",
            "Raw finger scan",
            "Previous month",
        ],
    )
    report_wb.calculation.calcMode = "auto"
    report_wb.calculation.fullCalcOnLoad = True
    report_wb.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_wb.save(output_path)
    log(f"Saved output: {output_path}")
    return output_path


def require_sheet(workbook: Workbook, name: str) -> Worksheet:
    sheet = find_sheet(workbook, name)
    if sheet is None:
        raise ValueError(f"ไม่พบชีทที่จำเป็น: {name}")
    return sheet


def find_sheet(workbook: Workbook, name: str) -> Worksheet | None:
    normalized_name = normalize_text(name)
    for sheet_name in workbook.sheetnames:
        if normalize_text(sheet_name) == normalized_name:
            return cast(Worksheet, workbook[sheet_name])
    return None


def clear_sheet(sheet: Worksheet) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    if sheet.max_column:
        sheet.delete_cols(1, sheet.max_column)


def is_readonly_merged_cell(sheet: Worksheet, row: int, column: int) -> bool:
    return isinstance(sheet.cell(row, column), MergedCell)


def writable_cell(sheet: Worksheet, row: int, column: int) -> Cell:
    cell = sheet.cell(row, column)
    if isinstance(cell, MergedCell):
        coordinate = cell.coordinate
        for merged_range in sheet.merged_cells.ranges:
            if coordinate in merged_range:
                return sheet.cell(merged_range.min_row, merged_range.min_col)
    return cast(Cell, cell)


def set_cell_value(sheet: Worksheet, row: int, column: int, value: object) -> Cell:
    cell = writable_cell(sheet, row, column)
    cell.value = value
    return cell


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\u200b\ufeff]", "", text)
    text = re.sub(r"[^0-9a-zA-Zก-๙]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_employee_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_header_map(sheet: Worksheet, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[header_row]:
        if cell.value is not None:
            headers[normalize_text(cell.value)] = cell.column
    return headers


def find_column(header_map: dict[str, int], candidates: Iterable[str], fallback: int | None = None) -> int | None:
    candidate_keys = [normalize_text(candidate) for candidate in candidates]
    for candidate in candidate_keys:
        if candidate in header_map:
            return header_map[candidate]
    for header, column in header_map.items():
        if any(candidate and candidate in header for candidate in candidate_keys):
            return column
    return fallback


def read_hr_employees(sheet: Worksheet) -> list[Employee]:
    header = read_header_map(sheet)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 1)
    id_col = find_column(header, ["หมายเลขพนักงานบนระบบ", "employee id", "emp id", "รหัสพนักงาน", "id"], 2)
    department_col = find_column(header, ["department", "แผนก", "ฝ่าย"], 3)

    employees: list[Employee] = []
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value if name_col else None
        if not name:
            continue
        employee_id = sheet.cell(row, id_col).value if id_col else ""
        department = sheet.cell(row, department_col).value if department_col else ""
        employees.append(
            Employee(
                name=str(name).strip(),
                normalized_name=normalize_text(name),
                employee_id=normalize_employee_id(employee_id),
                department=str(department or "").strip(),
            )
        )
    return employees


def read_finger_scan_rows(workbook: Workbook) -> list[dict[str, object]]:
    sheet = cast(
        Worksheet,
        workbook["info (used)"] if "info (used)" in workbook.sheetnames
        else workbook["info"] if "info" in workbook.sheetnames
        else workbook.active,
    )
    header_row = detect_header_row(sheet, ["department", "ชื่อ", "name", "เวลา", "date", "time"])
    header = read_header_map(sheet, header_row)

    dept_col = find_column(header, ["department", "แผนก", "ฝ่าย"], 1)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 2)
    id_col = find_column(header, ["หมายเลขพนักงานบนระบบ", "employee id", "emp id", "รหัสพนักงาน", "id"], 3)
    datetime_col = find_column(header, ["วันและเวลา", "date time", "datetime", "scan time", "เวลา", "date"], 4)

    rows: list[dict[str, object]] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        raw_datetime = sheet.cell(row, datetime_col).value if datetime_col else None
        raw_name = sheet.cell(row, name_col).value if name_col else None
        rows.append(
            {
                "source_department": sheet.cell(row, dept_col).value if dept_col else "",
                "raw_name": raw_name,
                "raw_employee_id": sheet.cell(row, id_col).value if id_col else "",
                "raw_datetime": raw_datetime,
            }
        )
    return rows


def detect_header_row(sheet: Worksheet, keywords: Iterable[str], max_scan_rows: int = 20) -> int:
    normalized_keywords = [normalize_text(keyword) for keyword in keywords]
    best_row = 1
    best_score = -1
    for row in range(1, min(sheet.max_row, max_scan_rows) + 1):
        values = [normalize_text(cell.value) for cell in sheet[row] if cell.value is not None]
        score = sum(1 for value in values for keyword in normalized_keywords if keyword and keyword in value)
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def read_previous_month_id_hints(sheet: Worksheet | None) -> dict[str, str]:
    if sheet is None:
        return {}
    header = read_header_map(sheet)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 1)
    id_col = find_column(header, ["หมายเลขพนักงานบนระบบ", "employee id", "emp id", "รหัสพนักงาน", "id"], 2)
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value if name_col else None
        employee_id = sheet.cell(row, id_col).value if id_col else None
        if name and employee_id:
            counts[normalize_text(name)][normalize_employee_id(employee_id)] += 1
    return {name: counter.most_common(1)[0][0] for name, counter in counts.items() if counter}


def read_previous_month_name_by_id(sheet: Worksheet | None) -> dict[str, str]:
    """Return mapping of normalized employee-ID → employee name from Previous Month sheet."""
    if sheet is None:
        return {}
    header = read_header_map(sheet)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 1)
    id_col = find_column(header, ["หมายเลขพนักงานบนระบบ", "employee id", "emp id", "รหัสพนักงาน", "id"], 2)
    mapping: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value if name_col else None
        employee_id = sheet.cell(row, id_col).value if id_col else None
        if name and employee_id:
            norm_id = normalize_employee_id(employee_id)
            if norm_id and norm_id not in mapping:
                mapping[norm_id] = str(name).strip()
    return mapping


def normalize_finger_rows(
    raw_rows: list[dict[str, object]],
    employees: list[Employee],
    id_hint_by_name: dict[str, str],
    options: ProcessingOptions,
    log: LogFn,
    name_by_id: dict[str, str] | None = None,
) -> list[FingerScanRow]:
    employees_by_normalized_name = {employee.normalized_name: employee for employee in employees}
    employee_names = list(employees_by_normalized_name.keys())
    raw_id_counts_by_matched_name: dict[str, Counter[str]] = defaultdict(Counter)
    interim: list[tuple[dict[str, object], Employee, date]] = []
    skipped_names: Counter[str] = Counter()
    skipped_invalid_dates: Counter[str] = Counter()

    for raw in raw_rows:
        if not raw.get("raw_name") or not raw.get("raw_datetime"):
            continue
        # If the raw_name looks like a pure ID, try to resolve it to a real name via Previous Month sheet
        raw_name_str = str(raw["raw_name"]).strip()
        if name_by_id and normalize_employee_id(raw_name_str) in name_by_id:
            raw = {**raw, "raw_name": name_by_id[normalize_employee_id(raw_name_str)]}
        employee = match_employee(raw["raw_name"], employees_by_normalized_name, employee_names, options.fuzzy_threshold)
        if employee is None:
            skipped_names[str(raw["raw_name"])] += 1
            continue
        scan_datetime = try_parse_scan_datetime(raw["raw_datetime"], options.year, options.month)
        if scan_datetime is None:
            skipped_invalid_dates[str(raw["raw_datetime"])] += 1
            continue
        if scan_datetime.month != options.month or scan_datetime.year != options.year:
            continue
        raw_employee_id = normalize_employee_id(raw.get("raw_employee_id"))
        if raw_employee_id:
            raw_id_counts_by_matched_name[employee.normalized_name][raw_employee_id] += 1
        interim.append((raw, employee, scan_datetime.date()))

    if skipped_names:
        total_skipped = sum(skipped_names.values())
        log(f"Warning: skipped {total_skipped} rows with no HR match ({len(skipped_names)} unique names):")
        for name, count in skipped_names.most_common():
            log(f"  - {name} ({count} rows)")
    if skipped_invalid_dates:
        total_skipped = sum(skipped_invalid_dates.values())
        examples = ", ".join(f"{value} ({count})" for value, count in skipped_invalid_dates.most_common(10))
        log(f"Warning: skipped {total_skipped} rows with invalid scan date/time. Top values: {examples}")

    resolved_id_by_name: dict[str, str] = {}
    for employee in employees:
        previous_hint = id_hint_by_name.get(employee.normalized_name)
        most_common_raw = ""
        if raw_id_counts_by_matched_name[employee.normalized_name]:
            most_common_raw = raw_id_counts_by_matched_name[employee.normalized_name].most_common(1)[0][0]
        resolved_id_by_name[employee.normalized_name] = previous_hint or most_common_raw or employee.employee_id

    return [
        FingerScanRow(
            source_department=str(raw.get("source_department") or "").strip(),
            raw_name=str(raw.get("raw_name") or "").strip(),
            matched_name=employee.name,
            raw_employee_id=normalize_employee_id(raw.get("raw_employee_id")),
            resolved_employee_id=resolved_id_by_name[employee.normalized_name],
            raw_datetime=raw.get("raw_datetime"),
            scan_date=scan_date,
            hr_department=employee.department,
        )
        for raw, employee, scan_date in interim
    ]


def match_employee(
    raw_name: object,
    employees_by_normalized_name: dict[str, Employee],
    employee_names: list[str],
    fuzzy_threshold: int,
) -> Employee | None:
    normalized = normalize_text(raw_name)
    if normalized in employees_by_normalized_name:
        return employees_by_normalized_name[normalized]
    match = process.extractOne(normalized, employee_names, scorer=fuzz.WRatio)
    if match and match[1] >= fuzzy_threshold:
        return employees_by_normalized_name[match[0]]
    return None


def parse_scan_datetime(value: object, default_year: int, target_month: int | None = None) -> datetime:
    if isinstance(value, datetime):
        if target_month and value.month != target_month and value.day == target_month:
            try:
                return value.replace(month=target_month, day=value.month)
            except ValueError:
                return value
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        base = datetime(1899, 12, 30)
        return base + timedelta(days=float(value))

    text = str(value).strip()
    has_thai_month = False
    for thai_month, month_number in THAI_MONTHS.items():
        if thai_month in text:
            text = text.replace(thai_month, str(month_number))
            has_thai_month = True
            break
    text = re.sub(r"เวลา|น\.", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()

    parsed = date_parser.parse(text, dayfirst=has_thai_month, fuzzy=True, default=datetime(default_year, 1, 1))
    if target_month and parsed.month != target_month and parsed.day == target_month:
        try:
            parsed = parsed.replace(month=target_month, day=parsed.month)
        except ValueError:
            pass
    year = parsed.year - 543 if parsed.year > 2400 else parsed.year
    return parsed.replace(year=year)


def try_parse_scan_datetime(value: object, default_year: int, target_month: int | None = None) -> datetime | None:
    try:
        return parse_scan_datetime(value, default_year, target_month)
    except (TypeError, ValueError, OverflowError):
        return None


def format_adjust_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def format_date_value(value: date) -> str:
    return f"{WEEKDAY_NAMES[value.weekday()]},{MONTH_NAMES[value.month - 1]} {value.day},{value.year}"


def write_raw_finger_scan_sheet(workbook: Workbook, rows: list[dict[str, object]], options: ProcessingOptions) -> None:
    if "Raw finger scan" in workbook.sheetnames:
        del workbook["Raw finger scan"]
    sheet = workbook.create_sheet("Raw finger scan")
    headers = [
        "Department",
        "ชื่อพนักงาน",
        "หมายเลขพนักงานบนระบบ",
        "วันและเวลา",
        "Adjust Date",
        "Date Value",
        "Department",
    ]
    sheet.append(headers)
    for index, row in enumerate(rows, start=2):
        raw_name = row.get("raw_name")
        raw_datetime = row.get("raw_datetime")
        if not raw_name and not raw_datetime:
            sheet.append([None, None, None, None, None, None, None])
            continue
        parsed_datetime: datetime | None = None
        if raw_datetime:
            parsed_datetime = try_parse_scan_datetime(raw_datetime, options.year, options.month)
        sheet.append(
            [
                row.get("source_department"),
                raw_name,
                row.get("raw_employee_id"),
                raw_datetime,
                datetime(parsed_datetime.year, parsed_datetime.month, parsed_datetime.day) if parsed_datetime else None,
                datetime(parsed_datetime.year, parsed_datetime.month, parsed_datetime.day) if parsed_datetime else None,
                f'=VLOOKUP(B{index},\'ฐานข้อมูลพนักงาน (HR)\'!C:D,2,0)',
            ]
        )
        cast(Any, sheet.cell(index, 5)).number_format = "dd/mm/yyyy"
        cast(Any, sheet.cell(index, 6)).number_format = "ddd,mmm d,yyyy"
    style_header(sheet)
    autofit_columns(sheet)


def write_count_scan_sheet(
    workbook: Workbook,
    employees: list[Employee],
    rows: list[dict[str, object]],
    id_hint_by_name: dict[str, str],
    options: ProcessingOptions,
) -> tuple[dict[str, int], int]:
    sheet = find_sheet(workbook, "Count Scan")
    existing_names: list[str] = []
    if sheet is not None:
        for row_index in range(5, sheet.max_row + 1):
            value = sheet.cell(row_index, 1).value
            if value and normalize_text(value) != "grand total":
                existing_names.append(str(value).strip())
    if sheet is None:
        sheet = workbook.create_sheet("Count Scan")
    else:
        clear_sheet(sheet)
    count_names = existing_names or [employee.name for employee in employees]
    scan_dates = sorted(get_scan_dates_for_month(rows, options))
    grand_total_col = 3 + len(scan_dates)
    total_col = grand_total_col + 1

    cast(Any, sheet.cell(2, 1)).value = "COUNTUNIQUE of ชื่อพนักงาน"
    cast(Any, sheet.cell(2, 2)).value = "Date Value"
    cast(Any, sheet.cell(3, 1)).value = "ชื่อพนักงาน"
    for column_index, scan_date in enumerate(scan_dates, start=3):
        date_cell = cast(Any, sheet.cell(3, column_index))
        date_cell.value = datetime(scan_date.year, scan_date.month, scan_date.day)
        date_cell.number_format = 'ddd", "mmm" "d", "yyyy'
    cast(Any, sheet.cell(3, grand_total_col)).value = "Grand Total"
    cast(Any, sheet.cell(3, total_col)).value = "Total"

    scans_by_name: dict[str, set[date]] = defaultdict(set)
    count_name_by_normalized = {normalize_text(name): name for name in count_names if name}
    normalized_count_names = list(count_name_by_normalized)
    count_name_by_employee_id = {
        employee_id: count_name_by_normalized[normalized_name]
        for normalized_name, employee_id in id_hint_by_name.items()
        if normalized_name in count_name_by_normalized and employee_id
    }
    for row in rows:
        if not row.get("raw_name") or not row.get("raw_datetime"):
            continue
        scan_datetime = try_parse_scan_datetime(row["raw_datetime"], options.year, options.month)
        if scan_datetime is None:
            continue
        if scan_datetime.year != options.year or scan_datetime.month != options.month:
            continue
        normalized_raw_name = normalize_text(row["raw_name"])
        raw_employee_id = normalize_employee_id(row.get("raw_employee_id"))
        count_name = count_name_by_employee_id.get(raw_employee_id) or count_name_by_normalized.get(normalized_raw_name)
        if count_name is None:
            match = process.extractOne(normalized_raw_name, normalized_count_names, scorer=fuzz.WRatio)
            if match and match[1] >= options.fuzzy_threshold:
                count_name = count_name_by_normalized[match[0]]
        if count_name:
            scans_by_name[normalize_text(count_name)].add(scan_datetime.date())

    totals: dict[str, int] = {}
    last_date_column = get_column_letter(grand_total_col - 1)
    date_totals = [0 for _ in scan_dates]
    for row_index, count_name in enumerate(count_names, start=5):
        employee_scan_dates = scans_by_name.get(normalize_text(count_name), set())
        day_values = [1.0 if scan_date in employee_scan_dates else None for scan_date in scan_dates]
        total = sum(1 for value in day_values if value == 1)
        totals[normalize_text(count_name)] = total
        cast(Any, sheet.cell(row_index, 1)).value = count_name
        for column_index, value in enumerate(day_values, start=3):
            cast(Any, sheet.cell(row_index, column_index)).value = value
            if value == 1.0:
                date_totals[column_index - 3] += 1
        cast(Any, sheet.cell(row_index, grand_total_col)).value = 1.0
        cast(Any, sheet.cell(row_index, total_col)).value = f"=sum(C{row_index}:{last_date_column}{row_index})"

    grand_total_row = len(count_names) + 5
    cast(Any, sheet.cell(grand_total_row, 1)).value = "Grand Total"
    cast(Any, sheet.cell(grand_total_row, 2)).value = 0.0
    for column_index, date_total in enumerate(date_totals, start=3):
        cast(Any, sheet.cell(grand_total_row, column_index)).value = float(date_total) if date_total else None
    cast(Any, sheet.cell(grand_total_row, grand_total_col)).value = float(len(count_names))
    cast(Any, sheet.cell(grand_total_row, total_col)).value = None

    for row_index in range(grand_total_row + 1, 1002):
        cast(Any, sheet.cell(row_index, total_col)).value = f"=sum(C{row_index}:{last_date_column}{row_index})"
    cast(Any, sheet.cell(1, total_col)).value = float(total_col)
    cast(Any, sheet.cell(4, 2)).value = 0.0
    cast(Any, sheet.cell(4, grand_total_col)).value = 0.0
    cast(Any, sheet.cell(4, total_col)).value = f"=sum(C4:{last_date_column}4)"

    style_header(sheet)
    autofit_columns(sheet)
    return totals, total_col


def get_scan_dates_for_month(rows: list[dict[str, object]], options: ProcessingOptions) -> set[date]:
    scan_dates: set[date] = set()
    for row in rows:
        if not row.get("raw_datetime"):
            continue
        scan_datetime = try_parse_scan_datetime(row["raw_datetime"], options.year, options.month)
        if scan_datetime is None:
            continue
        if scan_datetime.year == options.year and scan_datetime.month == options.month and scan_datetime.weekday() < 5:
            scan_dates.add(scan_datetime.date())
    return scan_dates


def iter_month_business_days(year: int, month: int) -> Iterable[date]:
    current = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    while current < end:
        if current.weekday() < 5:
            yield current
        current += timedelta(days=1)


def read_leave_totals(sheet: Worksheet, options: ProcessingOptions) -> dict[str, int]:
    header_row = detect_header_row(sheet, ["ชื่อ", "name", "วันลา", "leave", "date"])
    header = read_header_map(sheet, header_row)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 1)
    amount_col = find_leave_amount_column(header)
    date_col = find_leave_date_column(header)

    totals: dict[str, float] = defaultdict(float)
    for row in range(header_row + 1, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value if name_col else None
        if not name:
            continue
        if date_col:
            leave_date_raw = sheet.cell(row, date_col).value
            if leave_date_raw:
                try:
                    leave_date = parse_scan_datetime(leave_date_raw, options.year, options.month).date()
                    if leave_date.year != options.year or leave_date.month != options.month:
                        continue
                except Exception:
                    pass
        amount = sheet.cell(row, amount_col).value if amount_col else 1
        totals[normalize_text(name)] += parse_leave_amount(amount)
    return {name: math.ceil(total) for name, total in totals.items()}


def find_leave_amount_column(header: dict[str, int]) -> int | None:
    exact_candidates = [
        "total no of leave applied",
        "total no of leave",
        "total leave applied",
        "total leave",
        "จำนวนวันลา",
        "จำนวนวัน",
    ]
    for candidate in exact_candidates:
        column = header.get(normalize_text(candidate))
        if column:
            return column
    for title, column in header.items():
        if "total" in title and ("leave" in title or "day" in title):
            return column
    return find_column(header, ["amount", "days", "day"], None)


def find_leave_date_column(header: dict[str, int]) -> int | None:
    for candidate in ["fromdate", "from date", "วันที่เริ่ม", "วันที่"]:
        column = header.get(normalize_text(candidate))
        if column:
            return column
    return find_column(header, ["fromdate", "from date", "date", "วัน"], None)


def process_ptvn_report(
    ptvn_report_path: str | Path,
    ptvn_output_path: str | Path,
    source_report_sheet: Worksheet,
    employees: list[Employee],
    count_totals: dict[str, int],
    leave_totals: dict[str, int],
    options: ProcessingOptions,
) -> Path:
    ptvn_report_path = Path(ptvn_report_path)
    ptvn_output_path = Path(ptvn_output_path)
    workbook = load_workbook(ptvn_report_path, keep_vba=ptvn_report_path.suffix.lower() == ".xlsm")

    summary_sheet = require_sheet(workbook, "2026 Summary Report")
    individual_sheet = find_sheet(workbook, "2026 Individual Attendance Report") or find_sheet(workbook, "2026 Individual Attendance Repo")
    if individual_sheet is None:
        raise ValueError("ไม่พบชีทที่จำเป็น: 2026 Individual Attendance Report")

    employee_department_by_name = read_ptvn_employee_departments(individual_sheet)
    if not employee_department_by_name:
        employee_department_by_name = read_report_employee_departments(source_report_sheet, employees)
    update_ptvn_individual_report(individual_sheet, count_totals, leave_totals, options)
    update_ptvn_summary_report(summary_sheet, individual_sheet, employee_department_by_name, count_totals, leave_totals, options)
    write_ptvn_dashboard(
        workbook,
        summary_sheet,
        individual_sheet,
        options,
        employee_department_by_name,
        count_totals,
        leave_totals,
        use_vba_dashboard_charts=ptvn_report_path.suffix.lower() == ".xlsm",
        ptvn_source_path=ptvn_report_path,
    )

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    ptvn_output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(ptvn_output_path)
    return ptvn_output_path


def read_report_employee_departments(source_report_sheet: Worksheet, employees: list[Employee]) -> dict[str, str]:
    header_row = find_annual_report_header_row(source_report_sheet)
    header = read_header_map(source_report_sheet, header_row)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 2)
    department_col = find_column(header, ["org count", "department", "แผนก", "ฝ่าย"], None)
    fallback_by_name = {employee.normalized_name: employee.department for employee in employees}

    employee_department_by_name: dict[str, str] = {}
    for row in range(header_row + 1, source_report_sheet.max_row + 1):
        name = source_report_sheet.cell(row, name_col).value if name_col else None
        if not name:
            continue
        normalized_name = normalize_text(name)
        department = source_report_sheet.cell(row, department_col).value if department_col else None
        employee_department_by_name[normalized_name] = str(department or fallback_by_name.get(normalized_name, "")).strip()

    for employee in employees:
        employee_department_by_name.setdefault(employee.normalized_name, employee.department)
    return employee_department_by_name


def read_ptvn_employee_departments(individual_sheet: Worksheet) -> dict[str, str]:
    header_row = find_annual_report_header_row(individual_sheet)
    header = read_header_map(individual_sheet, header_row)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 2)
    department_col = find_ptvn_org_count_column(header)

    employee_department_by_name: dict[str, str] = {}
    if not department_col:
        return employee_department_by_name
    for row in range(header_row + 1, individual_sheet.max_row + 1):
        name = individual_sheet.cell(row, name_col).value if name_col else None
        department = individual_sheet.cell(row, department_col).value if department_col else None
        if name and department:
            employee_department_by_name[normalize_text(name)] = str(department).strip()
    return employee_department_by_name


def update_ptvn_summary_report(
    sheet: Worksheet,
    individual_sheet: Worksheet,
    employee_department_by_name: dict[str, str],
    count_totals: dict[str, int],
    leave_totals: dict[str, int],
    options: ProcessingOptions,
) -> None:
    month_col = find_summary_month_column(sheet, options)
    working_days = options.working_days or len(list(iter_month_business_days(options.year, options.month)))
    set_cell_value(sheet, 3, month_col, working_days)

    individual_name = quote_sheet_name(individual_sheet.title)
    individual_header_row = find_annual_report_header_row(individual_sheet)
    individual_header = read_header_map(individual_sheet, individual_header_row)
    individual_department_col = find_ptvn_org_count_column(individual_header)
    individual_month_target_col = find_month_metric_column(individual_sheet, individual_header_row, options, ["target", "%", "เป้า"])
    if individual_month_target_col is None:
        individual_month_target_col = find_column(individual_header, ["target", "%", "เป้า"], None)
    if individual_month_target_col is None:
        raise ValueError("ไม่พบ column Target ในชีท 2026 Individual Attendance Report")
    department_letter = get_column_letter(individual_department_col)
    target_letter = get_column_letter(individual_month_target_col)

    department_rows: list[int] = []
    overall_row: int | None = None
    for row in range(1, sheet.max_row + 1):
        department = sheet.cell(row, 2).value
        if not department:
            continue
        normalized_department = normalize_text(department)
        if "overall" in normalized_department:
            overall_row = row
            continue
        if normalized_department in {"target", "department", "departments", "แผนก"}:
            continue
        if is_readonly_merged_cell(sheet, row, 3) or is_readonly_merged_cell(sheet, row, month_col):
            continue

        department_rows.append(row)
        set_cell_value(sheet, row, 3, f'=COUNTIF({individual_name}!${department_letter}:${department_letter},B{row})')
        attendance_cell = set_cell_value(
            sheet,
            row,
            month_col,
            f'=IFERROR(ROUNDUP(AVERAGEIF({individual_name}!${department_letter}:${department_letter},B{row},{individual_name}!${target_letter}:${target_letter}),2),0)',
        )
        attendance_cell.number_format = "0%"

    if overall_row and department_rows:
        if not is_readonly_merged_cell(sheet, overall_row, 3):
            set_cell_value(sheet, overall_row, 3, f"=SUM(C{department_rows[0]}:C{department_rows[-1]})")
        if not is_readonly_merged_cell(sheet, overall_row, month_col):
            overall_cell = set_cell_value(
                sheet,
                overall_row,
                month_col,
                f"=AVERAGE({get_column_letter(month_col)}{department_rows[0]}:{get_column_letter(month_col)}{department_rows[-1]})",
            )
            overall_cell.number_format = "0%"


def update_ptvn_individual_report(
    sheet: Worksheet,
    count_totals: dict[str, int],
    leave_totals: dict[str, int],
    options: ProcessingOptions,
) -> None:
    header_row = find_annual_report_header_row(sheet)
    header = read_header_map(sheet, header_row)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 2)
    working_col = find_month_metric_column(sheet, header_row, options, ["working day", "working", "record", "วันทำงาน"])
    leave_col = find_month_metric_column(sheet, header_row, options, ["leave", "off-site", "off site", "ลา"])
    target_col = find_month_metric_column(sheet, header_row, options, ["target", "%", "เป้า"])

    if working_col is None:
        working_col = find_column(header, ["working day", "working", "record", "วันทำงาน"])
    if leave_col is None:
        leave_col = find_column(header, ["leave", "off-site", "off site", "ลา"])
    if target_col is None:
        target_col = find_column(header, ["target", "%", "เป้า"])
    if not working_col:
        raise ValueError("ไม่พบ column Working Day ในชีท 2026 Individual Attendance Report")

    working_days = options.working_days or len(list(iter_month_business_days(options.year, options.month)))
    set_cell_value(sheet, 3, working_col, working_days)
    if target_col:
        set_cell_value(sheet, 3, target_col, math.ceil(working_days * 0.6))

    working_day_reference = f"${get_column_letter(working_col)}$3"
    for row in range(header_row + 1, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value if name_col else None
        if not name:
            continue
        normalized_name = normalize_text(name)
        set_cell_value(sheet, row, working_col, count_totals.get(normalized_name, 0))
        if leave_col:
            set_cell_value(sheet, row, leave_col, leave_totals.get(normalized_name, 0))
        if target_col:
            working_letter = get_column_letter(working_col)
            leave_letter = get_column_letter(leave_col) if leave_col else working_letter
            target_cell = set_cell_value(
                sheet,
                row,
                target_col,
                f"=MIN(ROUNDUP(({working_letter}{row}+{leave_letter}{row})/{working_day_reference},2),100%)",
            )
            target_cell.number_format = "0%"


def round_up_percentage(value: float, digits: int = 2) -> float:
    factor = 10**digits
    return min(math.ceil(value * factor) / factor, 1.0)


def find_summary_month_column(sheet: Worksheet, options: ProcessingOptions) -> int:
    month_tokens = {
        str(options.month),
        f"{options.month:02d}",
        MONTH_NAMES[options.month - 1].lower(),
        date(options.year, options.month, 1).strftime("%B").lower(),
    }
    for row in range(1, min(sheet.max_row, 8) + 1):
        for column in range(1, max(sheet.max_column, 15) + 1):
            value = normalize_text(sheet.cell(row, column).value)
            if value and any(month_token_matches_value(token, value) for token in month_tokens):
                return column
    return options.month + 3


def month_token_matches_value(token: str, value: str) -> bool:
    if not token or not value:
        return False
    if token.isdigit():
        return token == value
    return token == value or token in value


def write_ptvn_dashboard(
    workbook: Workbook,
    summary_sheet: Worksheet,
    individual_sheet: Worksheet,
    options: ProcessingOptions,
    employee_department_by_name: dict[str, str],
    count_totals: dict[str, int],
    leave_totals: dict[str, int],
    use_vba_dashboard_charts: bool = False,
    ptvn_source_path: str | Path | None = None,
) -> None:
    dashboard_name = "Dashboard"
    if dashboard_name in workbook.sheetnames:
        del workbook[dashboard_name]
    dashboard = workbook.create_sheet(dashboard_name, 0)

    month_col = find_summary_month_column(summary_sheet, options)
    working_days = summary_sheet.cell(3, month_col).value or options.working_days or len(list(iter_month_business_days(options.year, options.month)))
    month_title = f"{MONTH_NAMES[options.month - 1]} {options.year}"
    department_rows, overall = collect_ptvn_summary_rows(summary_sheet, month_col)
    if not department_rows:
        department_rows, overall = collect_ptvn_summary_formula_rows(summary_sheet, month_col, employee_department_by_name)
    # Load data_only copy for reading cached cell values from prior months (formulas not evaluated by openpyxl)
    if ptvn_source_path:
        _data_wb = load_workbook(Path(ptvn_source_path), data_only=True)
        _summary_readonly = (
            _data_wb["2026 Summary Report"]
            if "2026 Summary Report" in _data_wb.sheetnames
            else summary_sheet
        )
    else:
        _summary_readonly = summary_sheet
    dashboard_months = collect_dashboard_months(_summary_readonly, options)
    # Inject current month dept attendance computed directly from individual_sheet (AVERAGEIF formulas not evaluated)
    _curr_dept_rows = _compute_dept_attendance_from_individual(individual_sheet, options)
    if _curr_dept_rows:
        _curr_month_label = f"{THAI_MONTH_NAMES[options.month - 1]} {options.year + 543}"
        _curr_wd = options.working_days or len(list(iter_month_business_days(options.year, options.month)))
        _curr_overall = round(sum(att for _, _, att in _curr_dept_rows) / len(_curr_dept_rows), 4) if _curr_dept_rows else 0.0
        dashboard_months = [
            (label, mn, wd_val, (_curr_dept_rows if mn == options.month else rows), oa)
            for label, mn, wd_val, rows, oa in dashboard_months
        ]
        if not any(mn == options.month for _, mn, _, _, _ in dashboard_months):
            dashboard_months.append((_curr_month_label, options.month, _curr_wd, _curr_dept_rows, _curr_overall))
    total_employees = overall[1] if overall else sum(row[1] for row in department_rows)
    overall_attendance = overall[2] if overall else round_up_percentage(sum(row[2] for row in department_rows) / len(department_rows)) if department_rows else 0
    departments_at_target = sum(1 for _, _, attendance in department_rows if attendance >= 0.6)
    below_target = len(department_rows) - departments_at_target

    dark_blue = "2F5597"
    navy = "7DA7D9"
    light_blue = "CFE2F3"
    panel_fill = "D3E1F1"
    green = "6AA84F"
    orange = "E69138"
    red = "CC0000"
    purple = "6D9EEB"
    white = "FFFFFF"
    border_color = "EDEDED"
    full_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    dashboard.sheet_view.showGridLines = False
    dashboard.freeze_panes = "C11"
    for row in range(1, 88):
        dashboard.row_dimensions[row].height = 21
    dashboard.row_dimensions[1].height = 34
    dashboard.row_dimensions[2].height = 24
    dashboard.row_dimensions[9].height = 8  # spacer between panel title and charts

    dashboard.merge_cells("C1:P2")
    title_cell = cast(Any, dashboard["C1"])
    title_cell.value = f'="PTVN Attendance Dashboard  ("&IFERROR(INDEX($AK$2:$AK$13,MATCH($A$6,$AH$2:$AH$13,0)),"{month_title}")&")"'
    title_cell.font = Font(size=22, bold=True, color=white)
    title_cell.fill = PatternFill("solid", fgColor=navy)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    employee_rows = collect_dashboard_employee_rows(workbook, employee_department_by_name, count_totals, leave_totals, int(working_days or 0))
    employee_dropdown_names = build_employee_dropdown_names(employee_rows)
    write_sidebar(dashboard, options, month_title, dashboard_months, department_rows, employee_rows, dark_blue, light_blue, full_border)
    write_dashboard_filter_model(dashboard, summary_sheet, individual_sheet, options, dashboard_months, department_rows, employee_rows, total_employees)
    write_dashboard_defined_names(workbook, dashboard, employee_dropdown_names)

    department_start_row = 14
    department_end_row = department_start_row + len(department_rows) - 1
    month_count = max(1, len(dashboard_months))
    month_end_column = 42 + month_count - 1
    month_end_letter = get_column_letter(month_end_column)
    department_range = f"$AN${department_start_row}:$AN${department_end_row}"
    employee_range = f"$AO${department_start_row}:$AO${department_end_row}"
    attendance_matrix = f"$AP${department_start_row}:${month_end_letter}${department_end_row}"
    month_range = f"$AH$2:$AH${len(dashboard_months) + 1}"
    month_index = f'MATCH($A$6,{month_range},0)'
    selected_department_attendance = f'IFERROR(INDEX({attendance_matrix},MATCH($A$12,{department_range},0),{month_index}),0)'
    overall_for_month = f'IFERROR(INDEX($AP$3:${month_end_letter}$3,1,{month_index}),0)'
    pass_for_month = f'IFERROR(SUMPRODUCT(--(INDEX({attendance_matrix},0,{month_index})>=0.6)),0)'
    below_for_month = f'IFERROR(SUMPRODUCT(--(INDEX({attendance_matrix},0,{month_index})<0.6)),0)'
    employee_end_row = max(2, len(employee_rows) + 1)
    employee_month_start_col = 62
    employee_month_end_col = employee_month_start_col + month_count - 1
    employee_month_end_letter = get_column_letter(employee_month_end_col)
    employee_month_matrix = f"$BJ$2:${employee_month_end_letter}${employee_end_row}"
    employee_filter = f'((($A$12="ทั้งหมด")+($BG$2:$BG${employee_end_row}=$BH$3)>0)*(($A$19="ทั้งหมด")+($BD$2:$BD${employee_end_row}=$A$19)>0))'
    selected_employee_attendance = f'IFERROR(INDEX({employee_month_matrix},MATCH($A$19,$BD$2:$BD${employee_end_row},0),{month_index}),0)'
    selected_attendance_for_filter = f'IF($A$19<>"ทั้งหมด",{selected_employee_attendance},IF($A$12="ทั้งหมด",{overall_for_month},{selected_department_attendance}))'
    pass_employee_formula = f'IFERROR(SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}>=0.6)),0)'
    below_employee_formula = f'IFERROR(SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}<0.6)),0)'

    cards = [
        (
            "C4:D6",
            "Total Employees",
            f'="Total Employees"&CHAR(10)&IF($A$19<>"ทั้งหมด",1,IF($A$12="ทั้งหมด",$AN$2,IFERROR(INDEX({employee_range},MATCH($A$12,{department_range},0)),0)))',
            "formula",
            dark_blue,
        ),
        ("E4:F6", "Working Days", f'="Working Days"&CHAR(10)&IFERROR(INDEX($AP$2:${month_end_letter}$2,1,{month_index}),0)', "formula", green),
        (
            "G4:H6",
            "PTV Overall",
            f'="PTV Overall"&CHAR(10)&TEXT({selected_attendance_for_filter},"0%")',
            "formula",
            green if overall_attendance >= 0.6 else orange,
        ),
        (
            "I4:J6",
            "Pass Departments",
            f'="Pass Departments"&CHAR(10)&IF($A$12="ทั้งหมด",{pass_for_month},IF({selected_department_attendance}>=0.6,1,0))',
            "formula",
            purple,
        ),
        (
            "K4:L6",
            "Below Target",
            f'="Below Target"&CHAR(10)&IF($A$12="ทั้งหมด",{below_for_month},IF({selected_department_attendance}<0.6,1,0))',
            "formula",
            red if below_target else green,
        ),
        ("M4:N6", "Pass Employee", f'="Pass Employee"&CHAR(10)&{pass_employee_formula}', "formula", green),
        ("O4:P6", "Below Target Employee", f'="Below Target Employee"&CHAR(10)&{below_employee_formula}', "formula", red),
    ]
    for cell_range, label, value, value_type, color in cards:
        write_dashboard_card(dashboard, cell_range, label, value, value_type, color, white)

    table_header_row = 62
    table_first_data_row = table_header_row + 1

    write_dashboard_panel_title(dashboard, f"C{table_header_row - 1}:F{table_header_row - 1}", "Department Detail - use filter arrows", dark_blue, panel_fill)
    headers = ["Department", "No. Employee", f"Attendance {MONTH_NAMES[options.month - 1]}", "Status"]
    for column, header in enumerate(headers, start=3):
        cell = cast(Any, dashboard.cell(table_header_row, column))
        cell.value = header
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")
        cell.border = full_border

    for row_index, (department, employees, attendance) in enumerate(department_rows, start=table_first_data_row):
        source_row = department_start_row + (row_index - table_first_data_row)
        if row_index == table_first_data_row:
            dashboard.cell(row_index, 3).value = f'=IF($A$12="ทั้งหมด",$AN${source_row},$A$12)'
            dashboard.cell(row_index, 4).value = f'=IF($A$12="ทั้งหมด",$AO${source_row},IFERROR(INDEX({employee_range},MATCH($A$12,{department_range},0)),0))'
            attendance_formula = f'=IF($A$12="ทั้งหมด",INDEX($AP${source_row}:${month_end_letter}${source_row},1,{month_index}),IFERROR(INDEX({attendance_matrix},MATCH($A$12,{department_range},0),{month_index}),0))'
        else:
            dashboard.cell(row_index, 3).value = f'=IF($A$12="ทั้งหมด",$AN${source_row},"")'
            dashboard.cell(row_index, 4).value = f'=IF($A$12="ทั้งหมด",$AO${source_row},"")'
            attendance_formula = f'=IF($A$12="ทั้งหมด",INDEX($AP${source_row}:${month_end_letter}${source_row},1,{month_index}),"")'
        attendance_cell = cast(Any, dashboard.cell(row_index, 5))
        attendance_cell.value = attendance_formula
        attendance_cell.number_format = "0%"
        status_cell = cast(Any, dashboard.cell(row_index, 6))
        status_cell.value = f'=IF(E{row_index}="","",IF(E{row_index}>=60%,"Pass","Below Target"))'
        status_cell.font = Font(color=green if attendance >= 0.6 else red, bold=True)
        for column in range(3, 7):
            cell = cast(Any, dashboard.cell(row_index, column))
            cell.border = full_border
            cell.alignment = Alignment(vertical="center")
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=panel_fill)

    if department_rows:
        table_end_row = table_header_row + len(department_rows)
        detail_table = Table(displayName="DashboardDepartmentTable", ref=f"C{table_header_row}:F{table_end_row}")
        detail_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        dashboard.add_table(detail_table)

    if department_rows:
        from collections import Counter as _EmpCounter
        _dept_emp_counts = _EmpCounter(dept for _, dept, _ in employee_rows)
        max_emp_per_dept = max(_dept_emp_counts.values()) if _dept_emp_counts else 0
        write_dashboard_chart_sources(
            dashboard,
            summary_sheet,
            individual_sheet,
            options,
            month_col,
            department_rows,
            departments_at_target,
            below_target,
            overall_attendance,
            len(dashboard_months),
            len(employee_rows),
            dashboard_months,
            max_emp_per_dept,
        )
        attendance_source_rows = max(len(department_rows), max_emp_per_dept)
        write_dashboard_panel_title(dashboard, "C8:G8", "Attendance Status Summary", dark_blue, panel_fill)
        write_dashboard_panel_title(dashboard, "H8:P8", "Monthly PTV Overall Trend", dark_blue, panel_fill)
        write_dashboard_panel_title(dashboard, "Q8:V8", "Attendance Distribution", dark_blue, panel_fill)
        write_dashboard_panel_title(dashboard, "C28:P28", "Department Performance Ranking", dark_blue, panel_fill)
        write_dashboard_panel_title(dashboard, "C48:P48", "YTD Average Attendance by Department", dark_blue, panel_fill)

        if not use_vba_dashboard_charts:
            status_donut = DoughnutChart()
            status_data = Reference(dashboard, min_col=135, min_row=14, max_row=15)
            status_labels = Reference(dashboard, min_col=134, min_row=14, max_row=15)
            status_donut.add_data(status_data)
            status_donut.set_categories(status_labels)
            status_donut.visible_cells_only = False
            status_donut.holeSize = 55
            status_donut.height = 12.6
            status_donut.width = 19.0
            if status_donut.legend:
                status_donut.legend.position = "r"
                status_donut.legend.overlay = True
                status_donut.legend.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode="edge", yMode="edge",
                        x=0.773, y=0.777, w=0.183, h=0.169,
                    )
                )
            if status_donut.series:
                from openpyxl.chart.series import DataPoint  # type: ignore[attr-defined]
                # Pass slice: RGB(180, 229, 161) = #B4E5A1
                pt_pass = DataPoint(idx=0, invertIfNegative=False)
                pt_pass.graphicalProperties.solidFill = "B4E5A1"
                pt_pass.graphicalProperties.line.solidFill = "B4E5A1"
                # Below Target slice: RGB(251, 168, 175) = #FBA8AF
                pt_below = DataPoint(idx=1, invertIfNegative=False)
                pt_below.graphicalProperties.solidFill = "FBA8AF"
                pt_below.graphicalProperties.line.solidFill = "FBA8AF"
                status_donut.series[0].dPt = [pt_pass, pt_below]
                status_donut.series[0].dLbls = DataLabelList()
                status_donut.series[0].dLbls.showPercent = True
                status_donut.series[0].dLbls.showSerName = False
                status_donut.series[0].dLbls.showCatName = False
                status_donut.series[0].dLbls.showVal = False
                status_donut.series[0].dLbls.showLegendKey = False
                status_donut.series[0].dLbls.showLeaderLines = True
                _set_dLbls_font_size(status_donut.series[0].dLbls, 16)
            dashboard.add_chart(status_donut, "C10")
        if not use_vba_dashboard_charts:
            # Status summary cards removed to free dashboard space
            # Metrics are already shown in KPI cards and chart labels
            pass

        monthly_source_rows = max(1, len(dashboard_months))

        if not use_vba_dashboard_charts:
            trend_chart = LineChart()
            trend_data = Reference(dashboard, min_col=138, max_col=139, min_row=13, max_row=13 + monthly_source_rows)
            trend_labels = Reference(dashboard, min_col=137, min_row=14, max_row=13 + monthly_source_rows)
            trend_chart.add_data(trend_data, titles_from_data=True)
            trend_chart.set_categories(trend_labels)
            trend_chart.visible_cells_only = False
            trend_chart.height = 12.6
            trend_chart.width = 29.9
            trend_chart.plot_area.layout = Layout(
                manualLayout=ManualLayout(
                    layoutTarget="inner", xMode="edge", yMode="edge",
                    x=0.0035, y=0.0084, w=0.9745, h=0.8993,
                )
            )
            trend_chart.y_axis.scaling.min = 0
            trend_chart.y_axis.scaling.max = 1
            trend_chart.y_axis.majorUnit = 0.2
            trend_chart.y_axis.numFmt = "0%"
            trend_chart.y_axis.delete = True
            trend_chart.x_axis.axPos = "b"
            trend_chart.x_axis.delete = False
            trend_chart.x_axis.tickLblPos = "nextTo"
            if trend_chart.legend:
                trend_chart.legend.position = "b"
                trend_chart.legend.overlay = False
            if trend_chart.series:
                trend_chart.series[0].marker.symbol = "circle"
                trend_chart.series[0].marker.size = 7
                trend_chart.series[0].smooth = True
                trend_chart.series[0].dLbls = DataLabelList()
                trend_chart.series[0].dLbls.showVal = True
                trend_chart.series[0].dLbls.showPercent = True
                trend_chart.series[0].dLbls.position = "t"
                trend_chart.series[0].dLbls.showSerName = False
                trend_chart.series[0].dLbls.showCatName = False
                trend_chart.series[0].dLbls.showLegendKey = False
                _set_dLbls_font_size(trend_chart.series[0].dLbls, 16)
                trend_chart.series[0].graphicalProperties.line.solidFill = "0070C0"
                trend_chart.series[0].graphicalProperties.line.width = 31750
            if len(trend_chart.series) > 1:
                trend_chart.series[1].marker.symbol = "none"
                trend_chart.series[1].smooth = False
                trend_chart.series[1].graphicalProperties.line.solidFill = "CC0000"
                trend_chart.series[1].graphicalProperties.line.width = 19050
                trend_chart.series[1].graphicalProperties.line.prstDash = "lgDash"
            trend_chart.legend = None
            dashboard.add_chart(trend_chart, "H10")
        if not use_vba_dashboard_charts:
            write_chart_callout(dashboard, "O10:P10", "Target 60%", "CC0000", "FFF5F5", full_border)

        if not use_vba_dashboard_charts:
            performance_chart = BarChart()
            performance_chart.type = "col"
            performance_chart.title = "Average Attendance Rate by Department (%)"
            performance_chart.x_axis.title = None
            performance_chart.y_axis.title = None
            perf_pass_data = Reference(dashboard, min_col=136, min_row=13, max_row=13 + attendance_source_rows)
            perf_fail_data = Reference(dashboard, min_col=145, min_row=13, max_row=13 + attendance_source_rows)
            performance_labels = Reference(dashboard, min_col=131, min_row=14, max_row=13 + attendance_source_rows)
            performance_chart.add_data(perf_pass_data, titles_from_data=True)  # series[0] = pass (green)
            performance_chart.add_data(perf_fail_data, titles_from_data=True)  # series[1] = fail (red)
            performance_chart.set_categories(performance_labels)
            performance_chart.varyColors = False
            performance_chart.y_axis.scaling.min = 0
            performance_chart.y_axis.scaling.max = 1
            performance_chart.y_axis.majorUnit = 0.2
            performance_chart.y_axis.numFmt = "0%"
            performance_chart.x_axis.delete = False
            performance_chart.x_axis.tickLblPos = "nextTo"
            performance_chart.visible_cells_only = False
            performance_chart.height = 14.3
            performance_chart.width = 49.5
            performance_chart.gapWidth = 219
            performance_chart.overlap = 100
            # Tell Excel to skip NA() values as gaps so only one colored bar shows per dept
            cast(Any, performance_chart).dispBlanksAs = "gap"  # type: ignore[attr-defined]
            if len(performance_chart.series) >= 2:
                performance_chart.series[0].graphicalProperties.solidFill = "B4E5A1"  # green pass
                performance_chart.series[0].graphicalProperties.line.solidFill = "B4E5A1"
                performance_chart.series[1].graphicalProperties.solidFill = "FBA8AF"  # red fail
                performance_chart.series[1].graphicalProperties.line.solidFill = "FBA8AF"
                for s in performance_chart.series:
                    s.dLbls = DataLabelList()
                    s.dLbls.showVal = False  # labels handled by invisible overlay below
                    s.dLbls.showCatName = False
                    s.dLbls.showSerName = False
                    s.dLbls.showLegendKey = False
            performance_chart.legend = None

            # Invisible line overlay — shows data labels from EB (actual value, no NA()) 
            from openpyxl.chart import LineChart as _PerfLabelChart
            perf_label_chart = _PerfLabelChart()
            cast(Any, perf_label_chart).dispBlanksAs = "gap"  # type: ignore[attr-defined]
            perf_label_ref = Reference(dashboard, min_col=132, min_row=13, max_row=13 + attendance_source_rows)
            perf_label_chart.add_data(perf_label_ref, titles_from_data=True)
            perf_label_chart.y_axis.axId = performance_chart.y_axis.axId
            perf_label_chart.x_axis.axId = performance_chart.x_axis.axId
            if perf_label_chart.series:
                perf_label_chart.series[0].graphicalProperties.line.solidFill = "FFFFFF"  # white = invisible
                perf_label_chart.series[0].marker.symbol = "none"
                perf_label_chart.series[0].smooth = False
                perf_label_chart.series[0].dLbls = DataLabelList()
                perf_label_chart.series[0].dLbls.showVal = True
                perf_label_chart.series[0].dLbls.showCatName = False
                perf_label_chart.series[0].dLbls.showSerName = False
                perf_label_chart.series[0].dLbls.showLegendKey = False
                _set_dLbls_font_size(perf_label_chart.series[0].dLbls, 16)
            performance_chart += perf_label_chart

            # Add target 60% as a red dashed line via combo chart (BarChart + LineChart)
            from openpyxl.chart import LineChart as _LineChart
            target_line_chart = _LineChart()
            cast(Any, target_line_chart).dispBlanksAs = "gap"  # type: ignore[attr-defined]
            target_ref = Reference(dashboard, min_col=133, min_row=13, max_row=13 + attendance_source_rows)
            target_line_chart.add_data(target_ref, titles_from_data=True)
            target_line_chart.y_axis.axId = performance_chart.y_axis.axId
            target_line_chart.x_axis.axId = performance_chart.x_axis.axId
            if target_line_chart.series:
                target_line_chart.series[0].graphicalProperties.line.solidFill = "FF0000"
                target_line_chart.series[0].graphicalProperties.line.width = 19050  # 1.5pt
                target_line_chart.series[0].graphicalProperties.line.prstDash = "lgDash"
                target_line_chart.series[0].marker.symbol = "none"
                target_line_chart.series[0].smooth = False
                target_line_chart.series[0].dLbls = DataLabelList()
                target_line_chart.series[0].dLbls.showVal = False
                target_line_chart.series[0].dLbls.showSerName = False
                target_line_chart.series[0].dLbls.showCatName = False
                target_line_chart.series[0].dLbls.showLegendKey = False
            performance_chart += target_line_chart
            dashboard.add_chart(performance_chart, "C29")

        if not use_vba_dashboard_charts:
            dist_chart = BarChart()
            dist_chart.type = "bar"
            dist_chart.varyColors = False
            dist_data = Reference(dashboard, min_col=141, min_row=13, max_row=18)
            dist_labels = Reference(dashboard, min_col=140, min_row=14, max_row=18)
            dist_chart.add_data(dist_data, titles_from_data=True)
            dist_chart.set_categories(dist_labels)
            dist_chart.visible_cells_only = False
            dist_chart.height = 12.6
            dist_chart.width = 27.0
            dist_chart.legend = None
            dist_chart.gapWidth = 55
            dist_chart.x_axis.delete = False
            dist_chart.x_axis.tickLblPos = "nextTo"
            if dist_chart.series:
                from openpyxl.chart.series import DataPoint  # type: ignore[attr-defined]
                # Monochrome blue: darkest=≥80% (best), lightest=<20% (worst)
                dist_blue = ["1F3864", "2F5597", "4472C4", "5B9BD5", "9DC3E6"]
                for i, hex_color in enumerate(dist_blue):
                    pt = DataPoint(idx=i, invertIfNegative=False)
                    pt.graphicalProperties.solidFill = hex_color
                    pt.graphicalProperties.line.solidFill = hex_color
                    dist_chart.series[0].dPt.append(pt)
                dist_chart.series[0].dLbls = DataLabelList()
                dist_chart.series[0].dLbls.showVal = True
                dist_chart.series[0].dLbls.showPercent = True
                dist_chart.series[0].dLbls.showCatName = False
                dist_chart.series[0].dLbls.showSerName = False
                dist_chart.series[0].dLbls.showLegendKey = False
                dist_chart.series[0].dLbls.position = "outEnd"
                _set_dLbls_font_size(dist_chart.series[0].dLbls, 16)
            dashboard.add_chart(dist_chart, "Q10")

        if not use_vba_dashboard_charts and department_rows:
            ytd_source_rows = len(department_rows)
            ytd_chart = BarChart()
            ytd_chart.type = "col"
            ytd_pass_data = Reference(dashboard, min_col=152, min_row=60, max_row=60 + ytd_source_rows)
            ytd_fail_data = Reference(dashboard, min_col=153, min_row=60, max_row=60 + ytd_source_rows)
            ytd_labels = Reference(dashboard, min_col=142, min_row=61, max_row=60 + ytd_source_rows)
            ytd_chart.add_data(ytd_pass_data, titles_from_data=True)  # series[0] = pass (green)
            ytd_chart.add_data(ytd_fail_data, titles_from_data=True)  # series[1] = fail (red)
            ytd_chart.set_categories(ytd_labels)
            ytd_chart.varyColors = False
            ytd_chart.y_axis.scaling.min = 0
            ytd_chart.y_axis.scaling.max = 1
            ytd_chart.y_axis.majorUnit = 0.2
            ytd_chart.y_axis.numFmt = "0%"
            ytd_chart.x_axis.delete = False
            ytd_chart.x_axis.tickLblPos = "nextTo"
            ytd_chart.visible_cells_only = False
            ytd_chart.height = 14.3
            ytd_chart.width = 49.5
            ytd_chart.gapWidth = 219
            ytd_chart.overlap = 100
            ytd_chart.legend = None
            cast(Any, ytd_chart).dispBlanksAs = "gap"  # type: ignore[attr-defined]
            if len(ytd_chart.series) >= 2:
                ytd_chart.series[0].graphicalProperties.solidFill = "B4E5A1"  # green pass
                ytd_chart.series[0].graphicalProperties.line.solidFill = "B4E5A1"
                ytd_chart.series[1].graphicalProperties.solidFill = "FBA8AF"  # red fail
                ytd_chart.series[1].graphicalProperties.line.solidFill = "FBA8AF"
                for s in ytd_chart.series:
                    s.dLbls = DataLabelList()
                    s.dLbls.showVal = False  # labels handled by invisible overlay below
                    s.dLbls.showCatName = False
                    s.dLbls.showSerName = False
                    s.dLbls.showLegendKey = False
            # Invisible line overlay — shows data labels from EM (actual YTD value, no NA())
            from openpyxl.chart import LineChart as _YtdLabelChart
            ytd_label_chart = _YtdLabelChart()
            cast(Any, ytd_label_chart).dispBlanksAs = "gap"  # type: ignore[attr-defined]
            ytd_label_ref = Reference(dashboard, min_col=143, min_row=60, max_row=60 + ytd_source_rows)
            ytd_label_chart.add_data(ytd_label_ref, titles_from_data=True)
            ytd_label_chart.y_axis.axId = ytd_chart.y_axis.axId
            ytd_label_chart.x_axis.axId = ytd_chart.x_axis.axId
            if ytd_label_chart.series:
                ytd_label_chart.series[0].graphicalProperties.line.solidFill = "FFFFFF"  # white = invisible
                ytd_label_chart.series[0].marker.symbol = "none"
                ytd_label_chart.series[0].smooth = False
                ytd_label_chart.series[0].dLbls = DataLabelList()
                ytd_label_chart.series[0].dLbls.showVal = True
                ytd_label_chart.series[0].dLbls.showCatName = False
                ytd_label_chart.series[0].dLbls.showSerName = False
                ytd_label_chart.series[0].dLbls.showLegendKey = False
                _set_dLbls_font_size(ytd_label_chart.series[0].dLbls, 16)
            ytd_chart += ytd_label_chart
            # Red dashed 60% target line
            from openpyxl.chart import LineChart as _YtdLineChart
            ytd_target_chart = _YtdLineChart()
            cast(Any, ytd_target_chart).dispBlanksAs = "gap"  # type: ignore[attr-defined]
            ytd_target_ref = Reference(dashboard, min_col=144, min_row=60, max_row=60 + ytd_source_rows)
            ytd_target_chart.add_data(ytd_target_ref, titles_from_data=True)
            ytd_target_chart.y_axis.axId = ytd_chart.y_axis.axId
            ytd_target_chart.x_axis.axId = ytd_chart.x_axis.axId
            if ytd_target_chart.series:
                ytd_target_chart.series[0].graphicalProperties.line.solidFill = "FF0000"
                ytd_target_chart.series[0].graphicalProperties.line.width = 19050
                ytd_target_chart.series[0].graphicalProperties.line.prstDash = "lgDash"
                ytd_target_chart.series[0].marker.symbol = "none"
                ytd_target_chart.series[0].smooth = False
                ytd_target_chart.series[0].dLbls = DataLabelList()
                ytd_target_chart.series[0].dLbls.showVal = False
                ytd_target_chart.series[0].dLbls.showSerName = False
                ytd_target_chart.series[0].dLbls.showCatName = False
                ytd_target_chart.series[0].dLbls.showLegendKey = False
            ytd_chart += ytd_target_chart
            dashboard.add_chart(ytd_chart, "C49")

    widths = {
        "A": 35,
        "C": 30,
        "D": 14,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 16,
        "R": 16,
        "S": 16,
        "T": 16,
        "U": 16,
        "V": 16,
        "W": 16,
        "X": 16,
    }
    for column, width in widths.items():
        dashboard.column_dimensions[column].width = width
    dashboard.column_dimensions["B"].hidden = True
    # Hide columns Y onward up to EA (chart source data starts at EA col 131).
    for column_index in range(25, 131):
        col_letter = get_column_letter(column_index)
        dashboard.column_dimensions[col_letter].width = 10
        dashboard.column_dimensions[col_letter].hidden = True


def write_sidebar(
    sheet: Worksheet,
    options: ProcessingOptions,
    month_title: str,
    dashboard_months: list[tuple[str, int, object, list[tuple[str, int, float]], float]],
    department_rows: list[tuple[str, int, float]],
    employee_rows: list[tuple[str, str, float]],
    dark_blue: str,
    light_blue: str,
    border: Border,
) -> None:
    for row in range(4, 39):
        for column in range(1, 3):
            cell = cast(Any, sheet.cell(row, column))
            cell.fill = PatternFill("solid", fgColor="F8FBFF")
            cell.border = border

    sheet.merge_cells("A4:B4")
    title_cell = cast(Any, sheet["A4"])
    title_cell.value = "ตัวกรองข้อมูล"
    title_cell.font = Font(size=13, bold=True, color=dark_blue)
    title_cell.fill = PatternFill("solid", fgColor="FFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = border

    month_values = [month_label for month_label, _, _, _, _ in dashboard_months]
    if not month_values:
        month_values = [f"{THAI_MONTH_NAMES[options.month - 1]} {options.year + 543}"]

    departments = [department for department, _, _ in department_rows]
    department_values = ["ทั้งหมด", *departments]
    employee_values = ["ทั้งหมด", *[name for name, _, _ in employee_rows]]

    english_month_values = [
        f"{MONTH_NAMES[month_number - 1]} {options.year}"
        for _, month_number, _, _, _ in dashboard_months
    ]
    if not english_month_values:
        english_month_values = [f"{MONTH_NAMES[options.month - 1]} {options.year}"]
    write_filter_source(sheet, 34, month_values)
    write_filter_source(sheet, 37, english_month_values)
    write_filter_source(sheet, 35, department_values)
    write_employee_filter_source_formulas(sheet, 36, len(employee_rows))

    write_filter_box(sheet, 5, 10, "เดือน", month_values, dark_blue, light_blue, border, source_col=34)
    sheet["A6"] = f"{THAI_MONTH_NAMES[options.month - 1]} {options.year + 543}"
    clear_filter_preview_rows(sheet, 7, 10, border)
    write_filter_box(sheet, 11, 17, "แผนก", department_values, dark_blue, light_blue, border, source_col=35)
    clear_filter_preview_rows(sheet, 13, 17, border)
    write_filter_box(sheet, 18, 38, "พนักงาน", employee_values, dark_blue, light_blue, border, source_col=36, validation_formula="=INDIRECT($BH$2)")
    write_employee_preview_formulas(sheet, 20, 38, len(employee_rows), border)


def write_filter_source(sheet: Worksheet, column: int, values: list[str]) -> None:
    unique_values = list(dict.fromkeys(value for value in values if value))
    for row_index, value in enumerate(unique_values, start=2):
        cast(Any, sheet.cell(row_index, column)).value = value


def write_employee_filter_source_formulas(sheet: Worksheet, column: int, employee_count: int) -> None:
    sheet.cell(2, column).value = "ทั้งหมด"


def clear_filter_preview_rows(sheet: Worksheet, start_row: int, end_row: int, border: Border) -> None:
    for row_index in range(start_row, end_row + 1):
        for column in range(1, 3):
            cell = cast(Any, sheet.cell(row_index, column))
            cell.value = None
            cell.fill = PatternFill("solid", fgColor="F8FBFF")
            cell.border = border


def write_filter_box(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    label: str,
    values: list[str],
    dark_blue: str,
    light_blue: str,
    border: Border,
    search: bool = False,
    source_col: int | None = None,
    validation_formula: str | None = None,
) -> None:
    header = cast(Any, sheet.cell(start_row, 1))
    header.value = label
    header.font = Font(size=12, bold=True, color=dark_blue)
    header.fill = PatternFill("solid", fgColor="FFFFFF")
    header.alignment = Alignment(horizontal="left", vertical="center")
    header.border = border
    header_side = cast(Any, sheet.cell(start_row, 2))
    header_side.fill = PatternFill("solid", fgColor="FFFFFF")
    header_side.border = border

    display_values = values or ["ทั้งหมด"]
    selected_row = start_row + 1
    selected_cell = cast(Any, sheet.cell(selected_row, 1))
    selected_cell.value = display_values[0]
    selected_cell.font = Font(size=11, bold=True, color="1F2937")
    selected_cell.fill = PatternFill("solid", fgColor=light_blue)
    selected_cell.alignment = Alignment(horizontal="left", vertical="center")
    selected_cell.border = border

    selected_side = cast(Any, sheet.cell(selected_row, 2))
    selected_side.value = "▼"
    selected_side.font = Font(size=10, bold=True, color=dark_blue)
    selected_side.fill = PatternFill("solid", fgColor=light_blue)
    selected_side.alignment = Alignment(horizontal="center", vertical="center")
    selected_side.border = border

    preview_values = display_values[1:] or display_values
    for offset, value in enumerate(preview_values, start=selected_row + 1):
        if offset > end_row:
            break
        cell = cast(Any, sheet.cell(offset, 1))
        cell.value = value
        cell.font = Font(size=11, color="1F2937")
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        side_cell = cast(Any, sheet.cell(offset, 2))
        side_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        side_cell.border = border

    if (source_col or validation_formula) and display_values:
        last_row = len(list(dict.fromkeys(value for value in display_values if value))) + 1
        source_letter = get_column_letter(source_col) if source_col else ""
        validation = DataValidation(
            type="list",
            formula1=validation_formula or f"=${source_letter}$2:${source_letter}${last_row}",
            allow_blank=False,
            showErrorMessage=True,
        )
        validation.error = "Please choose a value from the dropdown list."
        validation.errorTitle = "Invalid filter value"
        validation.prompt = "Select a filter value from this dropdown."
        validation.promptTitle = label
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(sheet.cell(selected_row, 1))


def write_dashboard_defined_names(workbook: Workbook, sheet: Worksheet, employee_dropdown_names: dict[str, list[str]]) -> None:
    sheet_name = quote_sheet_name(sheet.title)
    for column_offset, (range_name, employee_names) in enumerate(employee_dropdown_names.items(), start=90):
        for row_index, employee_name in enumerate(employee_names, start=2):
            sheet.cell(row_index, column_offset).value = employee_name
        column_letter = get_column_letter(column_offset)
        end_row = max(2, len(employee_names) + 1)
        workbook.defined_names.add(
            DefinedName(
                range_name,
                attr_text=f"{sheet_name}!${column_letter}$2:${column_letter}${end_row}",
            )
        )


def quote_sheet_name(sheet_name: str) -> str:
    return f"'{sheet_name.replace(chr(39), chr(39) * 2)}'"


def build_employee_dropdown_names(employee_rows: list[tuple[str, str, float]]) -> dict[str, list[str]]:
    ranges: dict[str, list[str]] = {"Emp_All": ["ทั้งหมด"]}
    for employee_name, department, _ in employee_rows:
        ranges["Emp_All"].append(employee_name)
        range_name = employee_range_name(department)
        ranges.setdefault(range_name, ["ทั้งหมด"]).append(employee_name)
    return ranges


def employee_range_name(department: object) -> str:
    key = normalize_dashboard_key(department)
    return f"Emp_{key}" if key else "Emp_Blank"


def write_employee_preview_formulas(sheet: Worksheet, start_row: int, end_row: int, employee_count: int, border: Border) -> None:
    if employee_count <= 0:
        return
    source_end_row = employee_count + 1
    for row_index in range(start_row, end_row + 1):
        cell = cast(Any, sheet.cell(row_index, 1))
        list_position = row_index - start_row + 1
        cell.value = f'=IFERROR(INDEX(INDIRECT($BH$2),{list_position}),"")'
        cell.font = Font(size=11, color="1F2937")
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        side_cell = cast(Any, sheet.cell(row_index, 2))
        side_cell.value = None
        side_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        side_cell.border = border


def write_dashboard_panel_title(sheet: Worksheet, cell_range: str, title: str, color: str, fill_color: str) -> None:
    sheet.merge_cells(cell_range)
    start_cell = cell_range.split(":", 1)[0]
    cell = cast(Any, sheet[start_cell])
    cell.value = title
    cell.font = Font(size=16, bold=True, color=color)
    panel_fill = PatternFill("solid", fgColor=fill_color)
    panel_alignment = Alignment(horizontal="center", vertical="center")
    panel_border = Border(
        left=Side(style="thin", color="DDE7F3"),
        right=Side(style="thin", color="DDE7F3"),
        top=Side(style="thin", color="DDE7F3"),
        bottom=Side(style="thin", color="DDE7F3"),
    )
    # Apply fill/alignment/border to ALL cells in range (openpyxl requires this for merged cells)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            c_cell = cast(Any, sheet.cell(r, c))
            c_cell.fill = panel_fill
            c_cell.alignment = panel_alignment
            c_cell.border = panel_border


def _set_dLbls_font_size(dLbls: object, pt: int) -> None:
    """Set data label font size using openpyxl RichText / CharacterProperties."""
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (
        RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties,
    )
    sz = pt * 100  # e.g. 16pt → 1600 (hundredths of a point)
    rpr = CharacterProperties(sz=sz, b=False)
    para = Paragraph(pPr=ParagraphProperties(defRPr=rpr), endParaRPr=rpr)
    txPr = RichText(bodyPr=RichTextProperties(), p=[para])
    cast(Any, dLbls).txPr = txPr  # type: ignore[attr-defined]


def write_chart_callout(
    sheet: Worksheet,
    cell_range: str,
    text: str,
    color: str,
    fill_color: str,
    border: Border,
) -> None:
    sheet.merge_cells(cell_range)
    start_cell = cell_range.split(":", 1)[0]
    cell = cast(Any, sheet[start_cell])
    cell.value = text
    cell.font = Font(size=9, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            styled_cell = cast(Any, sheet.cell(row, column))
            styled_cell.fill = PatternFill("solid", fgColor=fill_color)
            styled_cell.border = border


def write_dashboard_donut_center_note(
    sheet: Worksheet,
    cell_range: str,
    employee_formula_ref: str,
    color: str,
    fill_color: str,
    border: Border,
) -> None:
    sheet.merge_cells(cell_range)
    start_cell = cell_range.split(":", 1)[0]
    cell = cast(Any, sheet[start_cell])
    cell.value = f'=IFERROR({employee_formula_ref},"")&CHAR(10)&"Employees"'
    cell.font = Font(size=15, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            styled_cell = cast(Any, sheet.cell(row, column))
            styled_cell.fill = PatternFill("solid", fgColor=fill_color)
            styled_cell.border = border
            styled_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_status_summary_cards(
    sheet: Worksheet,
    cell_range: str,
    employee_formula_ref: str,
    dark_blue: str,
    green: str,
    orange: str,
    border: Border,
) -> None:
    for row in range(12, 22):
        for column in range(8, 10):
            cell = cast(Any, sheet.cell(row, column))
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells("H12:I14")
    total_card = cast(Any, sheet["H12"])
    total_card.value = f'=IFERROR({employee_formula_ref},"")&CHAR(10)&"Employees"'
    total_card.font = Font(size=15, bold=True, color=dark_blue)
    total_card.fill = PatternFill("solid", fgColor="F5F8FF")
    total_card.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_card.border = border

    sheet.merge_cells("H16:I18")
    pass_card = cast(Any, sheet["H16"])
    pass_card.value = '="Pass"&CHAR(10)&$EE$14&" ("&TEXT(IFERROR($EE$14/SUM($EE$14:$EE$15),0),"0%")&")"'
    pass_card.font = Font(size=12, bold=True, color=green)
    pass_card.fill = PatternFill("solid", fgColor="F3FAF0")
    pass_card.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pass_card.border = border

    sheet.merge_cells("H19:I21")
    below_card = cast(Any, sheet["H19"])
    below_card.value = '="Below Target"&CHAR(10)&$EE$15&" ("&TEXT(IFERROR($EE$15/SUM($EE$14:$EE$15),0),"0%")&")"'
    below_card.font = Font(size=12, bold=True, color=orange)
    below_card.fill = PatternFill("solid", fgColor="FFF6EC")
    below_card.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    below_card.border = border


def write_absence_no_category_message(sheet: Worksheet, cell_range: str, color: str, border: Border) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cell = cast(Any, sheet.cell(row, column))
            cell.value = None
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells(cell_range)
    start_cell = cell_range.split(":", 1)[0]
    message_cell = cast(Any, sheet[start_cell])
    message_cell.value = "No absence category data available\nTotal absence days are calculated from attendance gap."
    message_cell.font = Font(size=11, bold=True, color=color)
    message_cell.fill = PatternFill("solid", fgColor="FFFFFF")
    message_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    message_cell.border = border


def write_absence_overview(
    sheet: Worksheet,
    cell_range: str,
    border: Border,
    fill_color: str,
    dark_blue: str,
    green: str,
    orange: str,
    red: str,
) -> None:
    for row in range(29, 50):
        for column in range(10, 17):
            cell = cast(Any, sheet.cell(row, column))
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    sheet.merge_cells("J30:L32")
    total_card = cast(Any, sheet["J30"])
    total_card.value = '="Total Absence Days"&CHAR(10)&TEXT($EQ$40,"0.0")'
    total_card.font = Font(size=13, bold=True, color=dark_blue)
    total_card.fill = PatternFill("solid", fgColor="F5F8FF")
    total_card.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_card.border = border

    sheet.merge_cells("M30:P32")
    rate_card = cast(Any, sheet["M30"])
    rate_card.value = '="Absence Rate"&CHAR(10)&TEXT($EQ$41,"0.0%")'
    rate_card.font = Font(size=13, bold=True, color=red)
    rate_card.fill = PatternFill("solid", fgColor="FFF5F5")
    rate_card.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rate_card.border = border

    sheet.merge_cells("J34:M34")
    breakdown_title = cast(Any, sheet["J34"])
    breakdown_title.value = "Absence Breakdown"
    breakdown_title.font = Font(size=11, bold=True, color=dark_blue)
    breakdown_title.fill = PatternFill("solid", fgColor=fill_color)
    breakdown_title.alignment = Alignment(horizontal="center")

    sheet.merge_cells("N34:P34")
    top_title = cast(Any, sheet["N34"])
    top_title.value = "Top 5 Employees by Absence Days"
    top_title.font = Font(size=11, bold=True, color=dark_blue)
    top_title.fill = PatternFill("solid", fgColor=fill_color)
    top_title.alignment = Alignment(horizontal="center")

    for index, row in enumerate(range(35, 40), start=1):
        sheet.merge_cells(start_row=row, start_column=14, end_row=row, end_column=15)
        name_cell = cast(Any, sheet.cell(row, 14))
        days_cell = cast(Any, sheet.cell(row, 16))
        source_row = 39 + index
        name_cell.value = f'=IF($EM${source_row}="","",{index}&". "&$EM${source_row})'
        days_cell.value = f'=IF($EN${source_row}="","",TEXT($EN${source_row},"0.0")&" days")'
        name_cell.font = Font(size=9, bold=True, color=dark_blue)
        days_cell.font = Font(size=9, bold=True, color=green if index <= 2 else orange)
        name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        days_cell.alignment = Alignment(horizontal="center", vertical="center")
        name_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        days_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        name_cell.border = border
        days_cell.border = border


def write_monthly_trend_sources(
    sheet: Worksheet,
    summary_sheet: Worksheet,
    options: ProcessingOptions,
    selected_month_overall: float,
) -> list[tuple[str, int]]:
    overall_row = find_summary_overall_row(summary_sheet)
    month_columns = find_summary_month_columns(summary_sheet, options)
    if not month_columns:
        month_columns = [(MONTH_NAMES[options.month - 1], find_summary_month_column(summary_sheet, options))]

    summary_name = quote_sheet_name(summary_sheet.title)
    sheet["AC13"] = "Month"
    sheet["AD13"] = "PTV Overall"
    for row_index, (month_name, column) in enumerate(month_columns, start=14):
        sheet.cell(row_index, 29).value = month_name
        attendance_cell = cast(Any, sheet.cell(row_index, 30))
        if overall_row:
            attendance_cell.value = f"={summary_name}!{get_column_letter(column)}{overall_row}"
        else:
            attendance_cell.value = selected_month_overall if month_name == MONTH_NAMES[options.month - 1] else 0
        attendance_cell.number_format = "0%"
    return month_columns


def find_summary_overall_row(summary_sheet: Worksheet) -> int | None:
    for row in range(1, summary_sheet.max_row + 1):
        value = summary_sheet.cell(row, 2).value
        if value and "overall" in normalize_text(value):
            return row
    return None


def find_summary_month_columns(summary_sheet: Worksheet, options: ProcessingOptions) -> list[tuple[str, int]]:
    month_columns: list[tuple[str, int]] = []
    header_row = None
    for row in range(1, min(summary_sheet.max_row, 8) + 1):
        values = [normalize_text(summary_sheet.cell(row, column).value) for column in range(1, summary_sheet.max_column + 1)]
        if any(month.lower() in values for month in MONTH_NAMES):
            header_row = row
            break
    if header_row is None:
        return month_columns
    for column in range(1, summary_sheet.max_column + 1):
        header = normalize_text(summary_sheet.cell(header_row, column).value)
        for month_index, month_name in enumerate(MONTH_NAMES, start=1):
            if header == month_name.lower():
                overall_row = find_summary_overall_row(summary_sheet)
                if overall_row and summary_sheet.cell(overall_row, column).value is not None:
                    month_columns.append((month_name, column))
                elif month_index <= options.month:
                    month_columns.append((month_name, column))
                break
    selected_month_name = MONTH_NAMES[options.month - 1]
    if not any(month_name == selected_month_name for month_name, _ in month_columns):
        month_columns.append((selected_month_name, find_summary_month_column(summary_sheet, options)))
    return month_columns


def write_dashboard_chart_sources(
    sheet: Worksheet,
    summary_sheet: Worksheet,
    individual_sheet: Worksheet,
    options: ProcessingOptions,
    month_col: int,
    department_rows: list[tuple[str, int, float]],
    departments_at_target: int,
    below_target: int,
    selected_month_overall: float,
    month_count: int,
    employee_count: int,
    dashboard_months: list[tuple[str, int, object, list[tuple[str, int, float]], float]],
    max_emp_per_dept: int = 0,
) -> None:
    attendance_rows = sorted(department_rows, key=lambda item: item[2], reverse=True)
    department_start_row = 14
    department_end_row = department_start_row + len(department_rows) - 1
    month_end_column = 42 + max(1, month_count) - 1
    month_end_letter = get_column_letter(month_end_column)
    employee_month_start_col = 62
    employee_month_end_col = employee_month_start_col + max(1, month_count) - 1
    employee_month_end_letter = get_column_letter(employee_month_end_col)
    department_range = f"$AN${department_start_row}:$AN${department_end_row}"
    employee_range = f"$AO${department_start_row}:$AO${department_end_row}"
    attendance_matrix = f"$AP${department_start_row}:${month_end_letter}${department_end_row}"
    employee_month_matrix = f"$BJ$2:${employee_month_end_letter}${max(2, employee_count + 1)}"
    month_range = f"$AH$2:$AH${max(1, month_count) + 1}"
    month_index = f'MATCH($A$6,{month_range},0)'
    selected_attendance = f'IFERROR(INDEX({attendance_matrix},MATCH($A$12,{department_range},0),{month_index}),0)'
    selected_employees = f'IFERROR(INDEX({employee_range},MATCH($A$12,{department_range},0)),0)'
    employee_end_row = max(2, employee_count + 1)
    selected_working_days = f'IFERROR(INDEX($AP$2:${month_end_letter}$2,1,{month_index}),0)'
    employee_filter = f'((($A$12="ทั้งหมด")+($BG$2:$BG${employee_end_row}=$BH$3)>0)*(($A$19="ทั้งหมด")+($BD$2:$BD${employee_end_row}=$A$19)>0))'
    employee_count_formula = f'IF($A$19<>"ทั้งหมด",1,IF($A$12="ทั้งหมด",$AN$2,{selected_employees}))'
    total_absence_formula = f'IFERROR(SUMPRODUCT(--({employee_filter}),({selected_working_days})*(1-$BF$2:$BF${employee_end_row})),0)'
    absence_rate_formula = f'IFERROR({total_absence_formula}/(({selected_working_days})*({employee_count_formula})),0)'

    sheet["EA13"] = "Department / Employee"
    sheet["EB13"] = "Attendance"
    sheet["EC13"] = "Target 60%"
    # employee_month_matrix: BJ2:end_letter n  (monthly attendance per employee)
    emp_month_end_letter = get_column_letter(62 + max(1, month_count) - 1)
    # Detect individual sheet column structure for direct employee lookup (bypasses BG/BH month mismatch)
    indiv_title = quote_sheet_name(individual_sheet.title)
    indiv_header_row = find_annual_report_header_row(individual_sheet)
    indiv_header = read_header_map(individual_sheet, indiv_header_row)
    indiv_org_col = find_ptvn_org_count_column(indiv_header)
    indiv_name_col = find_column(indiv_header, ["name", "ชื่อ", "employee name", "ชื่อพนักงาน"], None)
    indiv_pct_col = find_month_metric_column(individual_sheet, indiv_header_row, options, ["target", "%", "เป้า"])
    if indiv_pct_col is None:
        indiv_pct_col = find_column(indiv_header, ["target", "%", "เป้า"], None)
    indiv_data_start = indiv_header_row + 1
    indiv_data_end = individual_sheet.max_row
    indiv_name_col_letter = get_column_letter(indiv_name_col) if indiv_name_col else "B"
    indiv_org_col_letter = get_column_letter(indiv_org_col) if indiv_org_col else "D"
    indiv_pct_col_letter = get_column_letter(indiv_pct_col) if indiv_pct_col else "V"
    indiv_row_offset = indiv_data_start - 1
    perf_source_rows = max(len(attendance_rows), max_emp_per_dept)
    for row_index in range(14, 14 + perf_source_rows):
        k = row_index - 13  # 1-based rank within the list
        dept_row = attendance_rows[k - 1] if k <= len(attendance_rows) else None
        dept_name = excel_escape(dept_row[0]) if dept_row else ""
        # In dept mode: show dept name or blank; in emp mode: AGGREGATE k-th employee in selected dept
        dept_label = f'"{dept_name}"' if dept_name else '""'
        # AGGREGATE on individual sheet: find k-th row where Org Count = $A$12
        agg = (
            f'AGGREGATE(15,6,'
            f'(ROW({indiv_title}!${indiv_org_col_letter}${indiv_data_start}:${indiv_org_col_letter}${indiv_data_end})-{indiv_row_offset})'
            f'/({indiv_title}!${indiv_org_col_letter}${indiv_data_start}:${indiv_org_col_letter}${indiv_data_end}=$A$12)'
            f',{k})+{indiv_row_offset}'
        )
        emp_label = f'IFERROR(INDEX({indiv_title}!${indiv_name_col_letter}:${indiv_name_col_letter},{agg}),"")'  
        label_cell = cast(Any, sheet.cell(row_index, 131))
        label_cell.value = f'=IF($A$12="ทั้งหมด",{dept_label},{emp_label})'
        # Attendance: dept attendance in all-dept mode; direct BJ matrix lookup in emp mode
        # (avoids BF which depends on MATCH($A$6,AH) that breaks when A6 is Thai format)
        if dept_name:
            dept_attend = f'IFERROR(INDEX({attendance_matrix},MATCH("{dept_name}",{department_range},0),{month_index}),"")'
        else:
            dept_attend = '""'
        emp_attend = f'IFERROR(INDEX({indiv_title}!${indiv_pct_col_letter}:${indiv_pct_col_letter},{agg}),"")'  
        attendance_cell = cast(Any, sheet.cell(row_index, 132))
        attendance_cell.value = f'=IF($A$12="ทั้งหมด",{dept_attend},{emp_attend})'
        attendance_cell.number_format = "0%"
        target_cell = cast(Any, sheet.cell(row_index, 133))
        target_cell.value = f'=IF(EB{row_index}="","",60%)'
        target_cell.number_format = "0%"

    # Pass/Fail split columns for dynamic bar coloring (EF=136 pass, EO=145 fail)
    cast(Any, sheet["EF13"]).value = "Pass (\u226560%)"
    cast(Any, sheet["EO13"]).value = "Fail (<60%)"
    for row_index in range(14, 14 + perf_source_rows):
        pass_cell = cast(Any, sheet.cell(row_index, 136))
        pass_cell.value = f'=IF(EB{row_index}="",NA(),IF(EB{row_index}>=0.6,EB{row_index},NA()))'
        pass_cell.number_format = "0%"
        fail_cell = cast(Any, sheet.cell(row_index, 145))
        fail_cell.value = f'=IF(EB{row_index}="",NA(),IF(EB{row_index}<0.6,EB{row_index},NA()))'
        fail_cell.number_format = "0%"

    sheet["ED13"] = "Status"
    sheet["EE13"] = "Employees"
    sheet["ED14"] = '="Pass"&CHAR(10)&EE14&" Employees ("&TEXT(IFERROR(EE14/SUM($EE$14:$EE$15),0),"0%")&")"'
    sheet["EE14"] = f'=IFERROR(SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}>=0.6)),0)'
    sheet["ED15"] = '="Below Target"&CHAR(10)&EE15&" Employees ("&TEXT(IFERROR(EE15/SUM($EE$14:$EE$15),0),"0%")&")"'
    sheet["EE15"] = f'=IFERROR(SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}<0.6)),0)'

    sheet["EG13"] = "Month"
    sheet["EH13"] = "PTV Trend"
    sheet["EI13"] = "Target 60%"
    for row_index in range(14, 14 + max(1, month_count)):
        month_position = row_index - 13
        month_number = dashboard_months[month_position - 1][1] if month_position <= len(dashboard_months) else month_position
        cast(Any, sheet.cell(row_index, 137)).value = MONTH_NAMES[month_number - 1] if 1 <= month_number <= 12 else f"Month {month_position}"
        attendance_cell = cast(Any, sheet.cell(row_index, 138))
        attendance_cell.value = (
            f'=IF($A$19<>"ทั้งหมด",IFERROR(INDEX({employee_month_matrix},MATCH($A$19,$BD$2:$BD${employee_end_row},0),{month_position}),0),'
            f'IF($A$12<>"ทั้งหมด",IFERROR(INDEX({attendance_matrix},MATCH($A$12,{department_range},0),{month_position}),0),'
            f'IFERROR(INDEX($AP$3:${month_end_letter}$3,1,{month_position}),0)))'
        )
        attendance_cell.number_format = "0%"
        target_cell = cast(Any, sheet.cell(row_index, 139))
        target_cell.value = "=60%"
        target_cell.number_format = "0%"

    sheet["EJ39"] = "Absence Type"
    sheet["EK39"] = "Days"
    absence_types = ["Other", "Sick Leave", "Personal Leave", "Vacation Leave", "Business Leave", "Unpaid Leave"]
    for row_index, absence_type in enumerate(absence_types, start=40):
        cast(Any, sheet.cell(row_index, 140)).value = absence_type
        days_cell = cast(Any, sheet.cell(row_index, 141))
        days_cell.value = f"={total_absence_formula}" if absence_type == "Other" else "=0"
        days_cell.number_format = "0.0"

    sheet["EM39"] = "Top Absent Employee"
    sheet["EN39"] = "Absent Days"
    sheet["GH1"] = "Filtered Absence Days"
    for employee_row in range(2, employee_end_row + 1):
        row_filter = f'((($A$12="ทั้งหมด")+($BG{employee_row}=$BH$3)>0)*(($A$19="ทั้งหมด")+($BD{employee_row}=$A$19)>0))'
        helper_cell = cast(Any, sheet.cell(employee_row, 190))
        helper_cell.value = f'=IF({row_filter},{selected_working_days}*(1-$BF{employee_row}),-1)'
        helper_cell.number_format = "0.0"
    absence_helper_range = f"$GH$2:$GH${employee_end_row}"
    for row_index in range(40, 45):
        rank = row_index - 39
        cast(Any, sheet.cell(row_index, 143)).value = f'=IFERROR(INDEX($BD$2:$BD${employee_end_row},MATCH(LARGE({absence_helper_range},{rank}),{absence_helper_range},0)),"")'
        days_cell = cast(Any, sheet.cell(row_index, 144))
        days_cell.value = f'=IF(EM{row_index}="","",MAX(0,LARGE({absence_helper_range},{rank})))'
        days_cell.number_format = "0.0"

    sheet["EP39"] = "Absence Metric"
    sheet["EQ39"] = "Value"
    sheet["EP40"] = "Total Absence Days"
    cast(Any, sheet["EQ40"]).value = f"={total_absence_formula}"
    sheet["EP41"] = "Absence Rate"
    rate_cell = cast(Any, sheet["EQ41"])
    rate_cell.value = f"={absence_rate_formula}"
    rate_cell.number_format = "0.0%"

    # Attendance distribution bucket data (visible cols EJ:EK rows 13-18 for distribution chart)
    sheet["EJ13"] = "Attendance Range"
    sheet["EK13"] = "Employees"
    dist_buckets = [
        ("\u226580%",  f'SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}>=0.8))'),
        ("60-79%",  f'SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}>=0.6),--($BF$2:$BF${employee_end_row}<0.8))'),
        ("40-59%",  f'SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}>=0.4),--($BF$2:$BF${employee_end_row}<0.6))'),
        ("20-39%",  f'SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}>=0.2),--($BF$2:$BF${employee_end_row}<0.4))'),
        ("<20%",    f'SUMPRODUCT(--({employee_filter}),--($BF$2:$BF${employee_end_row}<0.2))'),
    ]
    for i, (label, formula) in enumerate(dist_buckets):
        cast(Any, sheet.cell(14 + i, 140)).value = label
        count_cell = cast(Any, sheet.cell(14 + i, 141))
        count_cell.value = f'=IFERROR({formula},0)'

    # YTD average attendance per department (EL=142, EM=143, EN=144 starting row 60)
    dept_monthly_vals: dict[str, list[float]] = {dept: [] for dept, _, _ in department_rows}
    # Track which months already have data so current month isn't double-counted
    current_month_in_history: set[str] = set()
    for _, month_num, _, month_dept_rows, _ in dashboard_months:
        month_dict = {d: att for d, _, att in month_dept_rows}
        for dept, _, _ in department_rows:
            if dept in month_dict and month_dict[dept] > 0:
                dept_monthly_vals[dept].append(month_dict[dept])
                if month_num == options.month:
                    current_month_in_history.add(dept)
    # Merge current month from department_rows (freshly computed) for depts not yet in history
    current_month_dict = {dept: att for dept, _, att in department_rows}
    for dept, _, _ in department_rows:
        if dept not in current_month_in_history and current_month_dict.get(dept, 0) > 0:
            dept_monthly_vals[dept].append(current_month_dict[dept])
    ytd_sorted = sorted(
        ((dept, sum(vals) / len(vals) if vals else 0.0) for dept, vals in dept_monthly_vals.items()),
        key=lambda x: x[1], reverse=True,
    )
    cast(Any, sheet["EL60"]).value = "Department"
    cast(Any, sheet["EM60"]).value = "YTD Avg (%)"
    cast(Any, sheet["EN60"]).value = "Target 60%"
    for i, (dept_name, ytd_avg) in enumerate(ytd_sorted):
        row = 61 + i
        cast(Any, sheet.cell(row, 142)).value = dept_name
        ytd_val_cell = cast(Any, sheet.cell(row, 143))
        ytd_val_cell.value = round(ytd_avg, 4)
        ytd_val_cell.number_format = "0%"
        ytd_target_cell = cast(Any, sheet.cell(row, 144))
        ytd_target_cell.value = 0.6
        ytd_target_cell.number_format = "0%"

    # YTD pass/fail split columns (EV=152 pass, EW=153 fail)
    cast(Any, sheet["EV60"]).value = "YTD Pass (\u226560%)"
    cast(Any, sheet["EW60"]).value = "YTD Fail (<60%)"
    for i, (dept_name, ytd_avg) in enumerate(ytd_sorted):
        row = 61 + i
        ytd_pass_cell = cast(Any, sheet.cell(row, 152))
        ytd_pass_cell.value = f'=IF(EM{row}="",NA(),IF(EM{row}>=0.6,EM{row},NA()))'
        ytd_pass_cell.number_format = "0%"
        ytd_fail_cell = cast(Any, sheet.cell(row, 153))
        ytd_fail_cell.value = f'=IF(EM{row}="",NA(),IF(EM{row}<0.6,EM{row},NA()))'
        ytd_fail_cell.number_format = "0%"


def write_dashboard_filter_model(
    sheet: Worksheet,
    summary_sheet: Worksheet,
    individual_sheet: Worksheet,
    options: ProcessingOptions,
    dashboard_months: list[tuple[str, int, object, list[tuple[str, int, float]], float]],
    department_rows: list[tuple[str, int, float]],
    employee_rows: list[tuple[str, str, float]],
    total_employees: int,
) -> None:
    summary_name = quote_sheet_name(summary_sheet.title)
    individual_name = quote_sheet_name(individual_sheet.title)
    individual_header_row = find_annual_report_header_row(individual_sheet)
    individual_header = read_header_map(individual_sheet, individual_header_row)
    individual_name_col = find_column(individual_header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 2)
    individual_name_letter = get_column_letter(individual_name_col or 2)
    individual_target_col = find_month_metric_column(individual_sheet, individual_header_row, options, ["target", "%", "เป้า"])
    if individual_target_col is None:
        individual_target_col = find_column(individual_header, ["target", "%", "เป้า"], None)
    individual_target_letter = get_column_letter(individual_target_col or 1)
    sheet["AN1"] = "Selected Department Source"
    sheet["AN2"] = f"=COUNTA({individual_name}!${individual_name_letter}${individual_header_row + 1}:${individual_name_letter}{individual_sheet.max_row})"
    sheet["AN13"] = "Department"
    sheet["AO13"] = "Employees"
    attendance_by_month_department = {
        month_label: {department: attendance for department, _, attendance in rows}
        for month_label, _, _, rows, _ in dashboard_months
    }
    for column_offset, (month_label, month_number, working_days, _, overall_attendance) in enumerate(dashboard_months, start=42):
        sheet.cell(1, column_offset).value = month_label
        summary_month_col = find_summary_month_number_column(summary_sheet, month_number, options.year) or find_summary_month_column(summary_sheet, options)
        summary_overall_row = find_summary_overall_row(summary_sheet)
        sheet.cell(2, column_offset).value = f"={summary_name}!{get_column_letter(summary_month_col)}3"
        overall_cell = cast(Any, sheet.cell(3, column_offset))
        overall_cell.value = f"={summary_name}!{get_column_letter(summary_month_col)}{summary_overall_row}" if summary_overall_row else overall_attendance
        overall_cell.number_format = "0%"
        sheet.cell(13, column_offset).value = month_label
    for row_index, (department, employees, _) in enumerate(department_rows, start=14):
        summary_row = find_summary_department_row(summary_sheet, department)
        cast(Any, sheet.cell(row_index, 40)).value = department
        cast(Any, sheet.cell(row_index, 41)).value = f"={summary_name}!C{summary_row}" if summary_row else employees
        sheet.cell(row_index, 61).value = normalize_dashboard_key(department)
        for column_offset, (month_label, _, _, _, _) in enumerate(dashboard_months, start=42):
            summary_month_col = find_summary_month_number_column(summary_sheet, dashboard_months[column_offset - 42][1], options.year) or find_summary_month_column(summary_sheet, options)
            attendance_cell = cast(Any, sheet.cell(row_index, column_offset))
            attendance_cell.value = f"={summary_name}!{get_column_letter(summary_month_col)}{summary_row}" if summary_row else attendance_by_month_department.get(month_label, {}).get(department, 0)
            attendance_cell.number_format = "0%"

    sheet["BD1"] = "Employee"
    sheet["BE1"] = "Department"
    sheet["BF1"] = "Attendance"
    sheet["BG1"] = "Department Key"
    sheet["BH1"] = "Employee Dropdown Range"
    department_end_row = 14 + len(department_rows) - 1
    fallback_key_formula = 'LOWER(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(TRIM($A$12)," ",""),"&",""),"(",""),")",""),"-",""))'
    sheet["BH2"] = f'=IF($A$12="ทั้งหมด","Emp_All","Emp_"&IFERROR(INDEX($BI$14:$BI${department_end_row},MATCH($A$12,$AN$14:$AN${department_end_row},0)),{fallback_key_formula}))'
    sheet["BH3"] = f'=IF($A$12="ทั้งหมด","ทั้งหมด",IFERROR(INDEX($BI$14:$BI${department_end_row},MATCH($A$12,$AN$14:$AN${department_end_row},0)),{fallback_key_formula}))'
    sheet["BI13"] = "Department Key"
    for row_index, (employee_name, department, attendance) in enumerate(employee_rows, start=2):
        sheet.cell(row_index, 56).value = employee_name
        sheet.cell(row_index, 57).value = department
        attendance_cell = cast(Any, sheet.cell(row_index, 58))
        month_count = max(1, len(dashboard_months))
        employee_month_end_letter = get_column_letter(62 + month_count - 1)
        attendance_cell.value = f'=IFERROR(INDEX($BJ{row_index}:${employee_month_end_letter}{row_index},1,MATCH($A$6,$AH$2:$AH${month_count + 1},0)),0)'
        attendance_cell.number_format = "0%"
        sheet.cell(row_index, 59).value = normalize_dashboard_key(department)

    for column_offset, (month_label, month_number, _, _, _) in enumerate(dashboard_months, start=62):
        sheet.cell(1, column_offset).value = month_label
        month_options = ProcessingOptions(
            month=month_number,
            year=options.year,
            fuzzy_threshold=options.fuzzy_threshold,
            working_days=options.working_days,
        )
        month_target_col = find_month_metric_column(individual_sheet, individual_header_row, month_options, ["target", "%", "เป้า"])
        if month_target_col is None:
            month_target_col = individual_target_col
        month_target_letter = get_column_letter(month_target_col or individual_target_col or 1)
        for row_index, (employee_name, _, _) in enumerate(employee_rows, start=2):
            employee_attendance_cell = cast(Any, sheet.cell(row_index, column_offset))
            employee_attendance_cell.value = f'=IFERROR(INDEX({individual_name}!${month_target_letter}:${month_target_letter},MATCH(BD{row_index},{individual_name}!${individual_name_letter}:${individual_name_letter},0)),0)'
            employee_attendance_cell.number_format = "0%"


def excel_escape(value: object) -> str:
    return str(value).replace('"', '""')


def normalize_dashboard_key(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^0-9a-zA-Zก-๙]+", "", text)


def _compute_dept_attendance_from_individual(
    individual_sheet: Worksheet,
    options: ProcessingOptions,
) -> list[tuple[str, int, float]]:
    """Compute current-month dept attendance directly from individual sheet scan/leave counts."""
    hr = find_annual_report_header_row(individual_sheet)
    hmap = read_header_map(individual_sheet, hr)
    org_col = find_ptvn_org_count_column(hmap)
    rec_col = find_month_metric_column(individual_sheet, hr, options, ["record", "working day", "working"])
    leave_col = find_month_metric_column(individual_sheet, hr, options, ["leave", "ลา"])
    wd = options.working_days or len(list(iter_month_business_days(options.year, options.month)))
    if not rec_col or wd <= 0:
        return []
    from collections import defaultdict as _dd
    dept_vals: dict[str, list[float]] = _dd(list)
    for row in range(hr + 1, individual_sheet.max_row + 1):
        dept = individual_sheet.cell(row, org_col).value
        scans = individual_sheet.cell(row, rec_col).value
        leaves = individual_sheet.cell(row, leave_col).value if leave_col else 0
        if dept and isinstance(scans, (int, float)):
            pct = min(round((scans + float(leaves or 0)) / wd, 4), 1.0)
            dept_vals[str(dept)].append(pct)
    return [(dept, len(vals), round(sum(vals) / len(vals), 4)) for dept, vals in dept_vals.items()]


def collect_dashboard_months(
    summary_sheet: Worksheet,
    options: ProcessingOptions,
) -> list[tuple[str, int, object, list[tuple[str, int, float]], float]]:
    months: list[tuple[str, int, object, list[tuple[str, int, float]], float]] = []
    for month in range(1, options.month + 1):
        month_col = find_summary_month_number_column(summary_sheet, month, options.year)
        if not month_col:
            continue
        rows, overall = collect_ptvn_summary_rows(summary_sheet, month_col)
        working_days = summary_sheet.cell(3, month_col).value or len(list(iter_month_business_days(options.year, month)))
        overall_attendance = overall[2] if overall else round_up_percentage(sum(row[2] for row in rows) / len(rows)) if rows else 0
        month_label = f"{THAI_MONTH_NAMES[month - 1]} {options.year + 543}"
        months.append((month_label, month, working_days, rows, overall_attendance))
    if not months:
        month_col = find_summary_month_column(summary_sheet, options)
        rows, overall = collect_ptvn_summary_rows(summary_sheet, month_col)
        working_days = summary_sheet.cell(3, month_col).value or options.working_days or len(list(iter_month_business_days(options.year, options.month)))
        overall_attendance = overall[2] if overall else round_up_percentage(sum(row[2] for row in rows) / len(rows)) if rows else 0
        months.append((f"{THAI_MONTH_NAMES[options.month - 1]} {options.year + 543}", options.month, working_days, rows, overall_attendance))
    return months


def find_summary_month_number_column(summary_sheet: Worksheet, month: int, year: int) -> int | None:
    month_tokens = {
        str(month),
        f"{month:02d}",
        MONTH_NAMES[month - 1].lower(),
        date(year, month, 1).strftime("%B").lower(),
    }
    for row in range(1, min(summary_sheet.max_row, 8) + 1):
        for column in range(1, summary_sheet.max_column + 1):
            value = normalize_text(summary_sheet.cell(row, column).value)
            if value and any(month_token_matches_value(token, value) for token in month_tokens):
                return column
    return None


def collect_monthly_overall(summary_sheet: Worksheet) -> tuple[str, list[tuple[str, float]]]:
    overall_row = None
    for row in range(1, summary_sheet.max_row + 1):
        value = summary_sheet.cell(row, 2).value
        if value and "overall" in normalize_text(value):
            overall_row = row
            break
    if overall_row is None:
        return "Attendance", []

    results: list[tuple[str, float]] = []
    for month_name in MONTH_NAMES:
        month_col = None
        token = month_name.lower()
        for row in range(1, min(summary_sheet.max_row, 8) + 1):
            for column in range(1, summary_sheet.max_column + 1):
                header = normalize_text(summary_sheet.cell(row, column).value)
                if header == token or token in header:
                    month_col = column
                    break
            if month_col:
                break
        if not month_col:
            continue
        value = coerce_attendance_value(summary_sheet.cell(overall_row, month_col).value)
        if value is not None:
            results.append((month_name, value))
    return "PTV Overall", results


def ensure_selected_month_overall(
    monthly_overall: list[tuple[str, float]],
    options: ProcessingOptions,
    selected_month_overall: float,
) -> list[tuple[str, float]]:
    selected_month = MONTH_NAMES[options.month - 1]
    ensured: list[tuple[str, float]] = []
    found_selected_month = False
    for month_name, attendance in monthly_overall:
        if normalize_text(month_name) == normalize_text(selected_month):
            ensured.append((month_name, selected_month_overall))
            found_selected_month = True
        else:
            ensured.append((month_name, attendance))
    if not found_selected_month:
        ensured.append((selected_month, selected_month_overall))
    return ensured


def coerce_attendance_value(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("="):
            return None
        match = re.search(r"\d+(?:\.\d+)?", text)
        if not match:
            return None
        number = float(match.group(0))
        return number / 100 if "%" in text or number > 1 else number
    return None


def collect_dashboard_employee_rows(
    workbook: Workbook,
    employee_department_by_name: dict[str, str],
    count_totals: dict[str, int],
    leave_totals: dict[str, int],
    working_days: int,
) -> list[tuple[str, str, float]]:
    individual_sheet = find_sheet(workbook, "2026 Individual Attendance Report") or find_sheet(workbook, "2026 Individual Attendance Repo")
    if individual_sheet is None:
        return []
    header_row = find_annual_report_header_row(individual_sheet)
    header = read_header_map(individual_sheet, header_row)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 1)
    department_col = find_ptvn_org_count_column(header)
    if not name_col:
        return []
    rows: list[tuple[str, str, float]] = []
    divisor = working_days or 1
    for row in range(header_row + 1, individual_sheet.max_row + 1):
        name = individual_sheet.cell(row, name_col).value
        if not name:
            continue
        employee_name = str(name).strip()
        normalized_name = normalize_text(employee_name)
        department = individual_sheet.cell(row, department_col).value if department_col else ""
        resolved_department = str(department or employee_department_by_name.get(normalized_name, "")).strip()
        attendance = min((count_totals.get(normalized_name, 0) + leave_totals.get(normalized_name, 0)) / divisor, 1.0)
        rows.append((employee_name, resolved_department, attendance))
    return rows


def find_ptvn_org_count_column(header: dict[str, int]) -> int:
    return header.get(normalize_text("Org Count")) or find_column(header, ["org count"], None) or 4


def collect_ptvn_summary_rows(summary_sheet: Worksheet, month_col: int) -> tuple[list[tuple[str, int, float]], tuple[str, int, float] | None]:
    rows: list[tuple[str, int, float]] = []
    overall: tuple[str, int, float] | None = None
    for row in range(1, summary_sheet.max_row + 1):
        department = summary_sheet.cell(row, 2).value
        employees = summary_sheet.cell(row, 3).value
        attendance = summary_sheet.cell(row, month_col).value
        if not department or not isinstance(employees, (int, float)) or not isinstance(attendance, (int, float)):
            continue
        item = (str(department), int(employees), float(attendance))
        if "overall" in normalize_text(department):
            overall = item
        elif normalize_text(department) != "target":
            rows.append(item)
    return rows, overall


def collect_ptvn_summary_formula_rows(
    summary_sheet: Worksheet,
    month_col: int,
    employee_department_by_name: dict[str, str],
) -> tuple[list[tuple[str, int, float]], tuple[str, int, float] | None]:
    employees_by_department = Counter(normalize_text(department) for department in employee_department_by_name.values() if department)
    rows: list[tuple[str, int, float]] = []
    overall: tuple[str, int, float] | None = None
    for row in range(1, summary_sheet.max_row + 1):
        department = summary_sheet.cell(row, 2).value
        if not department:
            continue
        normalized_department = normalize_text(department)
        if "overall" in normalized_department:
            overall = (str(department), sum(employees_by_department.values()), 0.0)
        elif normalized_department != "target":
            count = employees_by_department.get(normalized_department, 0)
            if count:
                rows.append((str(department), count, 0.0))
    return rows, overall


def find_summary_department_row(summary_sheet: Worksheet, department: object) -> int | None:
    normalized_department = normalize_text(department)
    for row in range(1, summary_sheet.max_row + 1):
        if normalize_text(summary_sheet.cell(row, 2).value) == normalized_department:
            return row
    return None


def write_dashboard_card(
    sheet: Worksheet,
    cell_range: str,
    label: str,
    value: object,
    value_type: str,
    color: str,
    fill_color: str,
) -> None:
    sheet.merge_cells(cell_range)
    start_cell = cell_range.split(":", 1)[0]
    cell = cast(Any, sheet[start_cell])
    card_fills = {
        "Total Employees": "EFF6FF",
        "Working Days": "F3FAF0",
        "PTV Overall": "FFF6EC",
        "Pass Departments": "F4F1FF",
        "Below Target": "FFF1F1",
        "Pass Employee": "F3FAF0",
        "Below Target Employee": "FFF1F1",
    }
    fill = card_fills.get(label, fill_color)
    if value_type == "formula":
        display_value = str(value)
    elif value_type == "percent" and isinstance(value, (int, float)):
        display_value = f"{value:.0%}"
    elif value_type == "text":
        display_value = str(value)
    else:
        display_value = f"{value} {value_type}" if value_type not in {"days", "people", "departments"} else f"{value}"
    cell.value = display_value if value_type == "formula" else f"{label}\n{display_value}"
    cell.font = Font(size=16, bold=True, color=color)
    card_border = Border(
        left=Side(style="thin", color="DDE7F3"),
        right=Side(style="thin", color="DDE7F3"),
        top=Side(style="thin", color="DDE7F3"),
        bottom=Side(style="thin", color="DDE7F3"),
    )
    cell.border = card_border
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            styled_cell = cast(Any, sheet.cell(row, column))
            styled_cell.fill = PatternFill("solid", fgColor=fill)
            styled_cell.border = card_border
            styled_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_leave_amount(value: object) -> float:
    if value in (None, ""):
        return 1.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else 1.0


def update_annual_report(
    sheet: Worksheet,
    count_totals: dict[str, int],
    leave_totals: dict[str, int],
    options: ProcessingOptions,
    count_scan_total_col: int,
) -> None:
    header_row = find_annual_report_header_row(sheet)
    header = read_header_map(sheet, header_row)
    name_col = find_column(header, ["ชื่อพนักงาน", "employee name", "name", "ชื่อ"], 2)

    working_col = find_month_metric_column(sheet, header_row, options, ["working day", "working", "record", "วันทำงาน"])
    leave_col = find_month_metric_column(sheet, header_row, options, ["leave", "off-site", "off site", "ลา"])
    target_col = find_month_metric_column(sheet, header_row, options, ["target", "%", "เป้า"])

    if working_col is None:
        working_col = find_column(header, ["working day", "working", "record", "วันทำงาน"])
    if leave_col is None:
        leave_col = find_column(header, ["leave", "off-site", "off site", "ลา"])
    if target_col is None:
        target_col = find_column(header, ["target", "%", "เป้า"])

    if not working_col:
        raise ValueError("ไม่พบ column Working Day ในชีท 2026 Report")
    name_letter = get_column_letter(name_col) if name_col else "B"
    working_days = options.working_days or len(list(iter_month_business_days(options.year, options.month)))
    count_scan_total_letter = get_column_letter(count_scan_total_col)
    working_day_reference = f"${get_column_letter(working_col)}$3"

    set_cell_value(sheet, 3, working_col, working_days)
    if target_col:
        set_cell_value(sheet, 3, target_col, math.ceil(working_days * 0.6))

    for row in range(header_row + 1, sheet.max_row + 1):
        name = sheet.cell(row, name_col).value if name_col else None
        if not name:
            continue
        set_cell_value(
            sheet,
            row,
            working_col,
            f"=IFERROR(VLOOKUP({name_letter}{row},'Count Scan'!A:{count_scan_total_letter},{count_scan_total_col},0),0)",
        )
        if leave_col:
            set_cell_value(sheet, row, leave_col, f"=SUMIF('ข้อมูลวันลา'!B:B, {name_letter}{row}, 'ข้อมูลวันลา'!J:J)")
        if target_col:
            working_letter = get_column_letter(working_col)
            leave_letter = get_column_letter(leave_col) if leave_col else working_letter
            target_cell = set_cell_value(sheet, row, target_col, f"=({working_letter}{row}+{leave_letter}{row})/{working_day_reference}")
            target_cell.number_format = "0.00%"


def find_annual_report_header_row(sheet: Worksheet) -> int:
    for row in range(1, min(sheet.max_row, 20) + 1):
        values = [normalize_text(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
        has_name = any(value in {"name", "ชื่อ", "ชื่อพนักงาน", "employee name"} for value in values)
        has_record = any(value in {"record", "working day", "working"} for value in values)
        if has_name and has_record:
            return row
    return detect_header_row(sheet, ["name", "ชื่อ", "record", "working day", "leave", "target"])


def find_month_metric_column(
    sheet: Worksheet,
    header_row: int,
    options: ProcessingOptions,
    metric_keywords: Iterable[str],
) -> Optional[int]:
    month_tokens = {
        str(options.month),
        f"{options.month:02d}",
        MONTH_NAMES[options.month - 1].lower(),
        date(options.year, options.month, 1).strftime("%B").lower(),
    }
    for thai_month, month_number in THAI_MONTHS.items():
        if month_number == options.month:
            month_tokens.add(normalize_text(thai_month))
    metric_tokens = [normalize_text(keyword) for keyword in metric_keywords]
    all_month_tokens = {month.lower() for month in MONTH_NAMES}
    all_month_tokens.update(date(options.year, month, 1).strftime("%B").lower() for month in range(1, 13))
    all_month_tokens.update(normalize_text(thai_month) for thai_month in THAI_MONTHS)

    active_target_month = False
    for col in range(1, sheet.max_column + 1):
        month_header_values = [
            normalize_text(sheet.cell(row, col).value)
            for row in range(max(1, header_row - 3), header_row)
        ]
        month_header_text = " ".join(month_header_values)
        if any(token and token in month_header_text for token in month_tokens):
            active_target_month = True
        elif any(token and token in month_header_text for token in all_month_tokens):
            active_target_month = False

        metric_values = [normalize_text(sheet.cell(row, col).value) for row in range(max(1, header_row - 2), header_row + 1)]
        metric_text = " ".join(metric_values)
        has_metric = any(token and token in metric_text for token in metric_tokens)
        if active_target_month and has_metric:
            return col
    return None


def style_header(sheet: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"


def autofit_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(cast(int, column_cells[0].column))
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def ensure_sheet_order(workbook: Workbook, preferred_order: list[str]) -> None:
    ordered = []
    used_titles = set()
    for name in preferred_order:
        sheet = find_sheet(workbook, name)
        if sheet and sheet.title not in used_titles:
            ordered.append(sheet)
            used_titles.add(sheet.title)
    remaining = [sheet for sheet in workbook.worksheets if sheet.title not in used_titles]
    cast(Any, workbook)._sheets = ordered + remaining
